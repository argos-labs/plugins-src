#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excel`
====================================
.. moduleauthor:: Jerry Chae <mcchae@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel
"""
#
# Authors
# ===========
#
# * Jerry Chae, Kyobong Ahn
#
# Change Log
# --------
#
#  * [2023/06/22] Kyobong
#     - data only 사용시 xlwing에서 간헐적으로 sheet를 찾기 못하는 에러발생. sheet를 찾는 로직 추가
#     - "Sheet Finder" 기능 추가 기본값은 5
#  * [2023/05/25] Kyobong
#     - _erase_trailing_empty_values 에서 셀값이  0일경우 삭제하지 않도록 수정
#  * [2022/08/10] Kyobong
#     - password 기능 변경 tempdir에 저장 하는 방식 -> 비밀번호를 삭제 작업후에 다시 셋팅
#     - remove password 기능추가
#     - file, book(통합문서보호), sheet(시트보호), 각각 비밀번호 설정할수있는 기능 추가
#  * [2022/04/26] Kyobong
#     - write_cell에 오류가 있었음. 해당 문제 해결.
#  * [2022/03/23] Jerry
#     - 암호로 보호된 엑셀 파일인 경우 내부적으로 암호없는 것으로 저장 후 읽기, 임시파일은 삭제
#     - pywin32를 이용하므로 이 기능은 윈도우에서만 가능
#  * [2022/02/21] Kyobong
#     - 읽기만 할때  datetime에서 date만 출력 되게하는 Date 기능추가
#  * [2022/02/14] Kyobong
#     - 원본 데이터의 date_type을 그대로 유지하도록 수정.
#  * [2022/02/07] Kyobong
#     - dataonly 사용할때 백그라운드에 excel이 남아있는 버그가있슴. xlwing 사용할때 app.quit()로 닫아줌
#  * [2021/06/17]
#     - Data-Only로 읽어오는데 문제 디버깅 [by Shige]
#     - xlwings 로 읽도록 수정 (교봉씨 excelread 참조 및 수정)
#  * [2021/05/26]
#     - A{{rp.index}} 에 repeat를 돌면서 CSV에 넣는데 무조건 덮어 씀
#  * [2021/03/29]
#     - 그룹에 "2-Business Apps" 넣음
#  * [2021/01/14]
#     - --keep-blank의 표시이름 "Keep Ext Blanks" => "Keep Blanks"
#  * [2020/10/12]
#     - Write용 엑셀을 읽는데 30초 이상 걸리는 문제
#  * [2020/08/19]
#     - 만약 특정 파일명으로 출력하도록 되어 있는 경우, 출력을 이 파일명으로 지정
#     - --overwrite 옵션 추가하여 --data-only 여도 덮어쓰도록 함
#  * [2020/08/14]
#     - --set-cell, --set-value 이고 --data-only 인 경우 "파일명 (n).xlsx" 적용
#  * [2020/08/13]
#     - 다시 --with-formula 대신 --data-only 넣음
#     - 만약 --data-only 이고 출력을 동일 파일로 할 경우 저장된 formula가
#     -   모두 사라질 수 있으므로 "파일명 (n).xlsx" 식으로 변경
#  * [2020/08/05]
#     - add --with-formula instead of --data-only
#  * [2020/03/19]
#     - Change order of parameters
#  * [2020/02/02]
#     - add --encoding
#  * [2019/12/18]
#     - csv를 --big 으로 읽고 나중에 close, remove tempfile에 예외 발생 너머감
#     - --big인 경우, 100개만 처리하는 것 막고 모든 레코드 출력
#  * [2019/11/20]
#     - add new --find-string and --find-partial
#  * [2019/11/13]
#     - suppress warning which cause error from PAM
#  * [2019/08/19]
#     - "Copy of Bot Queue Demo.xlsx" 과 같이 헤더 또는 내용에 ,,,,,,, 가 포함되는 것을
#       제외 시킴
#     - --keep-blank 가 설정되어 있으면 위의 규칙을 안 따름
#     - add --set-cell, --set-value 로 입력 엑셀에 특정 셀 값 설정
#  * [2019/09/10]
#     - 결과의 앞뒤로 strip() 시켜서 출력
#  * [2019/07/19]
#     - add --dimensions
#  * [2019/07/18]
#     - add --clear-cell option
#  * [2019/04/25]
#     - set argument's displayname
#  * [2019/04/17]
#     - add csv read
#  * [2019/03/15]
#     - lineterminator='\n' # in windows twice newline
#  * [2018/11/28]
#     - starting
#
# TODO:
#   - display format, value ?
#   - 병합 cell 처리?

################################################################################
import os
import sys
import csv
import time
import shutil
# noinspection PyPackageRequirements
import xlwings
import openpyxl
import tempfile
import subprocess
from tempfile import gettempdir
# noinspection PyPackageRequirements
from openpyxl.utils import get_column_letter
# from openpyxl.workbook.protection import WorkbookProtection
# from openpyxl.worksheet.protection import SheetProtection
from io import StringIO
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import ModuleContext, func_log, \
    ArgsError, ArgsExit, get_icon_path
import warnings
warnings.filterwarnings("ignore")


################################################################################
# TODO: 목록에서 구하도록?
ENCODINGS = [
    "auto: Detecting",
    "cp932: Japanese",
    "eucjp: Japanese",
    "cp936: Chinese (simplified)",
    "cp949: Korean",
    "euckr: Korean",
    "cp950: Chinese (traditional)",
    "utf-8 UTF-8",
]


################################################################################
class Excel(object):
    # ==========================================================================
    def __init__(self, argspec, mcxt):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel file "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.opened = False
        self.wb = None
        self.ws = None
        self.min_row = self.max_row = -1
        self.min_col = self.max_col = -1
        self.rr = None
        # for csv reading
        self.tempfile = None

        # for logger
        self.mcxt = mcxt

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    def __repr__(self):
        sio = StringIO()
        wr = csv.writer(sio, lineterminator='\n')
        for row in self.rr:
            wr.writerow(row)
        return sio.getvalue()

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def xls2csv(self, filename, ws):
        with open(filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in ws.rows:
                c.writerow([cell.value for cell in r])
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.argspec.big and self.argspec.reverse:
            raise RuntimeError('--big and --reverse option cannot be set at the same time')
        if self.extension == '.csv':
            temp_name = next(tempfile._get_candidate_names())
            self.tempfile = os.path.join(gettempdir(), f'{temp_name}.xlsx')
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.load_workbook(self.filename,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if self.argspec.sheet is None:
                self.ws = self.wb.active
            elif self.argspec.sheet in self.wb.sheetnames:
                self.ws = self.wb[self.argspec.sheet]
            else:
                self.ws = self.wb.create_sheet(self.argspec.sheet)
        self.opened = True
        return self.opened

    # ==========================================================================
    def save_formula(self):
        if not (self.opened and self.argspec.formula):
            return False
        for fp in self.argspec.formula:
            fr, fm = fp.split('=', maxsplit=1)
            self.ws[fr].value = '=%s' % fm
        # 윈도우에서는 열었던 것을 저장할 수 없다함. 다른 파일로 임시로 저장하고
        # renmae 시켜봄
        tmp_fn = '%s-tmp.xlsx' % self.filename
        self.wb.save(tmp_fn)
        self.close()
        shutil.move(tmp_fn, self.filename)
        self.open(read_only=True)
        return True

    # ==========================================================================
    def list_sheet(self):
        if not self.opened:
            return ''
        # noinspection PyProtectedMember
        sl = [x.title for x in self.wb._sheets]
        return ','.join(sl)

    # ==========================================================================
    def close(self):
        if self.opened:
            self.wb.close()
            if self.tempfile and os.path.exists(self.tempfile) and \
                    self.tempfile != self.argspec.write:
                # wb를 close 후에 삭제하여야 함
                for _ in range(3):
                    # noinspection PyBroadException
                    try:
                        os.remove(self.tempfile)
                        break
                    except Exception:
                        time.sleep(0.3)
            self.ws = None
            self.wb = None
            self.opened = False
            return True
        return False

    # ==========================================================================
    def calc_range(self):
        if not self.opened:
            return False
        if not self.argspec.range:
            self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
            self.min_col, self.max_col = self.ws.min_column, self.ws.max_column
        else:
            r1, r2 = self.argspec.range.strip().split(':')
            c1, c2 = self.ws[r1.strip()], self.ws[r2.strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row, c2.row)
            self.min_col, self.max_col = min(c1.column, c2.column), max(c1.column, c2.column)
        return True

    # ==========================================================================
    def _get(self):
        # for r in range(self.min_row, self.max_row+1):
        #     row = []
        #     for c in range(self.min_col, self.max_col+1):
        #         cl = get_column_letter(c)
        #         v = self.ws['%s%d' % (cl, r)].value
        #         if v is None:
        #             v = ''
        #         row.append(v)
        #     yield row
        ws1, wbxl = None, None
        if self.argspec.data_only:
            app = xlwings.App(visible=False)
            wbxl = app.books.open(self.filename)
            if self.argspec.sheet:
                for i in range(self.argspec.sheetfind):
                    try:
                        ws1 = wbxl.sheets[self.argspec.sheet]
                        # sheet를 정상적으로 불러왔는지 확인. 간헐적으로 가져오지못함.
                        if ws1.range('A1').value:
                            break
                    except Exception as err:
                        # print(f'retry {i}')
                        self.mcxt.logger.error(f'get sheet retry {i+1}')
                        time.sleep(0.5)
                        continue
            else:
                ws1 = wbxl.sheets[0]

        for r in range(self.min_row, self.max_row + 1):
            row = []
            for c in range(self.min_col, self.max_col + 1):
                cl = get_column_letter(c)
                if self.argspec.data_only:
                    if str(type(ws1['%s%d' % (cl, r)].value)) == "<class 'float'>":
                        v = ws1['%s%d' % (cl, r)].value
                        if v - int(v) == 0:
                            v = int(v)
                    else:
                        if self.argspec.date and str(type(ws1['%s%d' % (cl, r)].value)) == "<class 'datetime.datetime'>":
                            v = ws1['B5'].value.date()
                        else:
                            v = ws1['%s%d' % (cl, r)].value
                elif self.argspec.write:
                    v = self.ws['%s%d' % (cl, r)]
                else:
                    if self.argspec.date and self.ws['%s%d' % (cl, r)].data_type == 'd':
                        v = self.ws['B5'].value.date()
                    else:
                        v = self.ws['%s%d' % (cl, r)].value
                if v is None:
                    v = ''
                row.append(v)
            yield row
        if self.argspec.data_only:
            wbxl.close()
            app.quit()

    # ==========================================================================
    def _erase_trailing_empty_values(self):
        rr = self.rr
        dl = []
        max_col = 0
        for i, r in enumerate(rr):
            for j in range(len(r)-1, -1, -1):
                # if r[j]:
                if r[j] or r[j] == 0:
                    if max_col < j:
                        max_col = j
                    break
                del r[j]
            if not r:
                dl.append(i)
        for i in dl[::-1]:
            del rr[i]
        for r in rr:
            for j in range(len(r), max_col+1):
                r.append('')

    # ==========================================================================
    def get(self):
        self.rr = []
        if self.argspec.big:
            wr = csv.writer(sys.stdout, lineterminator='\n')
            # 매우 큰 엑셀인 경우, 모든 Row를 처리하기 힘들기 때문에
            # 해당 Row 를 개별 stdout으로 출력
            for i, row in enumerate(self._get()):
                # 너무 커 질 수 있으므로 단지 처음 100 행만 저장
                # if i < 100:
                self.rr.append(row)
                wr.writerow(row)
                if 0 < i % 10 == 0:
                    sys.stdout.flush()
            sys.stdout.flush()
            return bool(self.rr)
        # not big option
        for row in self._get():
            self.rr.append(row)
        if not self.argspec.keep_blank:
            self._erase_trailing_empty_values()
        if self.argspec.reverse:    # transpose row, column
            self.rr = list(map(list, zip(*self.rr)))
        # s = str(self)
        s = str(self).strip()  # 셀 하나인 경우에도 개행문자가 붙는 것 제거
        # 2020/08/19: 특정 파일명으로 출력하도록 되어 있는 경우 제외
        if not self.argspec.write:
            sys.stdout.write(s)
        return bool(self.rr)

    # ==========================================================================
    def find(self, find_str, is_partial=False):
        find_cells = list()
        for r in range(self.min_row, self.max_row+1):
            for c in range(self.min_col, self.max_col+1):
                cl = get_column_letter(c)
                v = self.ws['%s%d' % (cl, r)].value
                if v is None:
                    v = ''
                is_find = False
                if is_partial:
                    if str(v).find(find_str) >= 0:
                        is_find = True
                else:
                    if str(v) == find_str:
                        is_find = True
                if is_find:
                    find_cells.append('%s%d' % (cl, r))
        return find_cells

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn, overwrite=False):
        if overwrite:
            return fn
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def save(self):
        argspec_write = self.argspec.write
        # 만약 쓸 결과가 없으면 False 리턴
        if not (self.opened and self.rr):
            return False
        # 만약 쓸 파일이 무효하다면 False 리턴
        if not (argspec_write and isinstance(argspec_write, str)):
            return False

        # 엑셀과 동일하게 출력하게 위하여 수정
        # - A{{rp.index}} 에 repeat를 돌면서 CSV에 넣는데 무조건 덮어 씀
        # 만약 결과를 csv에 쓴다면
        is_csv_write = False
        # 만약 이미 결과 CSV가 있거나 시작 위치가 처음 A1 이 아니라면
        if argspec_write.lower().endswith('csv') and \
                (os.path.exists(argspec_write) or self.argspec.write_cell != 'A1'):
            is_csv_write = True
        # 만약 CSV 이면 덮어 씀 (TODO: range write 적용 필요)
        if argspec_write.lower().endswith('csv'):  # and not is_csv_write:
            # with open(argspec_write, 'wb', newline='') as ofp:
            with open(argspec_write, 'w', newline='') as ofp:
                wr = csv.writer(ofp, lineterminator='\n')
                for row in self.rr:
                    wr.writerow(row)
            print(os.path.abspath(argspec_write), end='')
            return 0
        # 막약 출력할 엑셀 파일이 미리 존재하고 data_only 이면
        # 출력파일에 "파일명 (n).xlsx" 로 이름 변경 : 20200813
        if os.path.exists(argspec_write) and self.argspec.data_only:
            argspec_write = self._get_safe_next_filename(
                argspec_write, overwrite=self.argspec.overwrite)
            if self.argspec.write != argspec_write:
                shutil.copy(self.argspec.write, argspec_write)
        # 나머지는 모두 excel 파일이라 생각
        if os.path.exists(argspec_write):
            # 덮어쓰기용 파일에도 비밀번호가 걸려있는 경우. 단 읽은 엑셀일과 이름이 동일하면 안됨
            if self.argspec.write_password and self.filename != argspec_write:
                argspec_write = save_as_excel(argspec_write, self.argspec.write_password)
            w_wb = openpyxl.load_workbook(argspec_write)
        else:
            w_wb = openpyxl.Workbook()
        if self.argspec.write_sheet is None:
            w_ws = w_wb.active
        elif self.argspec.write_sheet in w_wb.sheetnames:
            w_ws = w_wb[self.argspec.write_sheet]
        else:
            w_ws = w_wb.create_sheet(self.argspec.write_sheet)
        s_cell = w_ws[self.argspec.write_cell]
        s_row, s_col = s_cell.row, s_cell.column
        for i, row in enumerate(self.rr):
            for j, v in enumerate(row):
                # w_ws[s_row+i][s_col+j].value = v
                try:
                    w_ws.cell(row=s_row+i, column=s_col+j, value=v.value)
                    w_ws[s_row+i][s_col+j-1].value = v.value
                    w_ws[s_row+i][s_col+j-1].data_type = v.data_type
                    w_ws[s_row+i][s_col+j-1].number_format = v.number_format
                    w_ws[s_row+i][s_col+j-1].quotePrefix = v.quotePrefix
                except:
                    w_ws.cell(row=s_row + i, column=s_col + j, value=v)
                # w_ws.cell(row=s_row+i, column=s_col+j, value=v)
        # set password
        # 통합문서 보호
        if self.argspec.set_pw_book:
            w_wb.security.workbookPassword = self.argspec.set_pw_book
            w_wb.security.lockStructure = True
        # 시트 보호
        if self.argspec.set_pw_sheet:
            w_ws.protection.sheet = True
            w_ws.protection.password = self.argspec.set_pw_sheet

        w_wb.save(argspec_write)
        print(os.path.abspath(argspec_write), end='')
        return 0

    # ==========================================================================
    def set_pw(self):
        if self.argspec.write:
            excel_file_path = self.argspec.write
        else:
            excel_file_path = self.filename
            # set password
            # 통합문서 보호
            if self.argspec.set_pw_book:
                self.wb.security.workbookPassword = self.argspec.set_pw_book
                self.wb.security.lockStructure = True
            # 시트 보호
            if self.argspec.set_pw_sheet:
                self.ws.protection.sheet = True
                self.ws.protection.password = self.argspec.set_pw_sheet
            # 둘중하나라도 설정되었다면 다시 저장
            if self.argspec.set_pw_book or self.argspec.set_pw_sheet:
                self.wb.save(self.filename)

        self.set_password(excel_file_path, self.argspec.set_pw_file)

    # ==========================================================================
    @staticmethod
    def set_password(excel_file_path, pw):

        from pathlib import Path

        excel_file_path = Path(excel_file_path)

        vbs_script = \
            f"""' Save with password required upon opening

        Set excel_object = CreateObject("Excel.Application")
        Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

        excel_object.DisplayAlerts = False
        excel_object.Visible = False

        workbook.SaveAs "{excel_file_path}",, "{pw}"

        excel_object.Application.Quit
        """

        # write
        vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
        with open(vbs_script_path, "w") as file:
            file.write(vbs_script)

        # execute
        subprocess.call(['cscript.exe', str(vbs_script_path)])

        # remove
        vbs_script_path.unlink()

        return

    # ==========================================================================
    def clear_cell(self):
        if not self.opened:
            return 1
        if not self.argspec.range:
            if self.extension == '.csv':
                os.remove(self.filename)
                with open(self.filename, 'w') as ofp:
                    ofp.write('')
            else:
                idx = self.wb.sheetnames.index(self.ws.title)
                self.wb.remove_sheet(self.ws)
                self.wb.create_sheet(self.ws.title, idx)
        else:
            for row in self.ws[self.argspec.range]:
                for cell in row:
                    cell.value = None
            if self.extension == '.csv':
                self.xls2csv(self.filename, self.ws)
        if self.extension != '.csv':
            self.wb.save(self.filename)
        return 0

    # ==========================================================================
    def set_cell(self, org_exl_f):
        if not self.opened:
            return 1
        self.filename = org_exl_f
        cell = self.ws[self.argspec.set_cell]
        cell.value = self.argspec.set_value
        if self.extension == '.csv':
            self.xls2csv(self.filename, self.ws)
        if self.extension != '.csv':
            filename_n = self.filename
            if self.argspec.data_only:
                filename_n = self._get_safe_next_filename(
                    self.filename, overwrite=self.argspec.overwrite)
            self.wb.save(filename_n)
        return 0


################################################################################
def save_as_excel(filepath, password):
    if not os.path.exists(filepath):
        raise IOError(f'Cannot read file "{filepath}"')
    bn = os.path.basename(filepath)
    fn, ext = os.path.splitext(bn)
    if not ext.lower().startswith('.xls'):
        raise RuntimeError(f'Password option is applied only for Excel file')
    if sys.platform != 'win32':
        raise RuntimeError(f'Password protected Excel file is only working on Windows OS')
    import win32com.client
    xl_file = os.path.abspath(filepath)
    sa_file = os.path.join(gettempdir(), bn)
    if os.path.exists(sa_file):
        os.remove(sa_file)
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        book = excel.Workbooks.Open(xl_file, False, False, None, password)
        excel.DisplayAlerts = False
        book.SaveAs(xl_file, None, '')
        book.Close()
        return xl_file
    except Exception as err:
        raise err


################################################################################
@func_log
def do_excel(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    org_exl_f = argspec.filename
    exl = None
    try:
        if argspec.reverse and argspec.big:
            raise RuntimeError('Cannot set Pivot and Big-Data Option cannot set together.')
        if argspec.password:
            argspec.filename = save_as_excel(argspec.filename, argspec.password)
        setattr(argspec, 'formula', None)
        exl = Excel(argspec, mcxt)
        data_only = argspec.data_only
        if exl.extension in ('.xlsx', '.csv'):
            exl.open(read_only=argspec.big, data_only=data_only)  # --big 이면 read_only로 처리
        elif exl.extension == '.xlsm':
            exl.open(read_only=False, data_only=data_only, keep_vba=True)
        else:
            raise RuntimeError('Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        exl.save_formula()
        if argspec.list_sheet:
            print(exl.list_sheet())
        elif argspec.dimensions:
            print(exl.ws.dimensions)
        elif argspec.find_string:
            exl.calc_range()
            r = exl.find(argspec.find_string, argspec.find_partial)
            print('\n'.join(r), end='')
        else:
            if exl.calc_range():
                exl.get()
                # print(str(exl))
                # 큰 데이터인 경우 _get yield로 row 별로 계속 stdout 으로 출력
            exl.save()
            if argspec.clear_cell:
                exl.clear_cell()
            if argspec.set_cell:
                exl.set_cell(org_exl_f)
        # 비밀번호 다시설정.
        if argspec.password and not argspec.remove_password and argspec.filename != argspec.write:
            exl.set_password(argspec.filename, argspec.password)
        if argspec.write_password and not argspec.remove_password:
            exl.set_password(argspec.write, argspec.write_password)
        if argspec.set_pw_file:
            exl.set_pw()

        sys.stdout.flush()
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    finally:
        if exl is not None:
            exl.close()
        # if pass_sa_file:
        #     os.remove(pass_sa_file)
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
        owner='ARGOS-LABS',
        group='2',   # Business Apps
        version='1.0',
        platform=['windows', 'darwin', 'linux'],
        output_type='csv',
        display_name='Excel Advanced',
        icon_path=get_icon_path(__file__),
        description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ##################################### for app dependent options
        # ----------------------------------------------------------------------
        mcxt.add_argument('--sheet', '-s', nargs='?', default=None,
                          display_name='Read-fr sheet', show_default=True,
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', '-r', nargs='?',
                          display_name='Read-fr range', show_default=True,
                          help='If set read the range (eg, "A3:C9") to handle otherwise get all sheet [[A3:C9]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--write', '-w', nargs='?', default=None,
                          show_default=True,
                          display_name='Write-to Excel file',
                          input_group='writing', input_method='filewrite',
                          help='Excel or CSV file to write, extension can be a xlsx, csv')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--write-sheet', nargs='?', default='Sheet1',
                          show_default=True,
                          display_name='Write-to sheet',
                          input_group='writing',
                          help='Sheet name to write. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--write-cell', nargs='?', default='A1',
                          show_default=True,
                          display_name='Write start cell',
                          const='A1',
                          input_group='writing',
                          help='Excel starting cell to write, default is [[A1]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--set-value',
                          display_name='Overwrite Value',
                          help='Overwrite cell Value positioned with "Overwirte Cell" for input excel')
        mcxt.add_argument('--set-cell',
                          display_name='Overwrite Cell',
                          help='Overwrite Cell Position to set with "Set Value" [[A1]] for input excel')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--find-string',
                          display_name='Find String',
                          help='Find string and get one or more cells')
        mcxt.add_argument('--find-partial',
                          display_name='Partial Match', action='store_true',
                          help='If this flag is set finding cells partially matching string')
        # ----------------------------------------------------------------------
        # mcxt.add_argument('--blank-col',
        #                   display_name='1st Black in Col',
        #                   help="Find first black cell in a Column like 'A'")
        mcxt.add_argument('--dimensions', action='store_true',
                          display_name='Show Data Range',
                          help="Display selected Sheet's dimensions like 'A1:C47'")
        # ----------------------------------------------------------------------
        # TODO: mutually_exclusive_group 인 경우, display_name은 오류 발생
        # ag = mcxt.add_mutually_exclusive_group()
        mcxt.add_argument('--reverse', action='store_true',
                          display_name='Pivot',
                          help='If set reversed row, column is chosen for reading')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--big', action='store_true',
                          display_name='Handle as Big-Data',
                          help='If set instead read all data into memory process line by line. Cannot use with --reverse option')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--clear-cell', action='store_true',
                          display_name='Clear Cell?',
                          help='If this flag is set then clear cells in range')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--keep-blank', action='store_true',
                          display_name='Keep Blanks',
                          help='')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          display_name='Encoding',
                          help='Encoding for CSV file')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--list-sheet', '-l', action='store_true',
                          display_name='List sheetnames only',
                          help='list Sheet names')
        mcxt.add_argument('--data-only', display_name='Data Only',
                          action='store_true',
                          help='If this flag is set get data only without formula')
        mcxt.add_argument('--overwrite', display_name='Allow Overwrite',
                          action='store_true',
                          help='If "Data Only" and this flag is set then overwrite without formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--date', display_name='Date',
                          action='store_true',
                          help='Shows datetime type values only as date.')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--remove-password', display_name='Remove Password',
                          action='store_true',
                          help='Remove password from excel file')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--password', display_name='Excel/CSV File',
                          input_group='Password',
                          input_method='password',
                          help='If Excel is protected by password then use this. This is only works on Excel installed Windows')
        mcxt.add_argument('--write-password', display_name='Write-to Excel file',
                          input_group='Password',
                          input_method='password',
                          help='If Excel is protected by password then use this. This is only works on Excel installed Windows')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--set-pw-file', display_name='File Password', default=None,
                          input_method='password',
                          input_group='Set password',
                          help='Set Password Protecting for Excel file')
        mcxt.add_argument('--set-pw-book', display_name='Book Password', default=None,
                          input_method='password',
                          input_group='Set password',
                          help='Set Password Protecting for Excel book')
        mcxt.add_argument('--set-pw-sheet', display_name='Sheet Password', default=None,
                          input_method='password',
                          input_group='Set password',
                          help='Set Password Protecting for Excel sheet')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--sheetfind', default=5, type=int,
                          display_name='Sheet Finder',
                          help='Data Only error solution.'
                               ' Attempts to find the selected sheet the specified number of times. default is [[5]]')
        # ----------------------------------------------------------------------
        # TODO: formula를 설정하는 것을 테스트 하는데,
        #     openpyxl 을 이용하여 formula를 설정하는 것은 가능하지만
        #     엑셀에서 다시 열어서 저장해야만 해당 값이 별도로 반영되는 문제
        #     formulas 등등의 모듈을 확인해 봤는데도 문제 있었음. 해결 결과의
        #     formula를 가져오는 것은 가능
        # mcxt.add_argument('--formula', action='append', default=None,
        #                   input_group='formula',
        #                   help='Excel formula cell. eg) --formula C54=SUM("C4:C53")')

        # ##################################### for app dependent parameters
        # ----------------------------------------------------------------------
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')
        argspec = mcxt.parse_args(args)
        return do_excel(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
