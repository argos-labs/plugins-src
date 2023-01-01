import openpyxl

filename = 'sample.xlsx'
wb = openpyxl.load_workbook(filename, read_only=True)
ws = wb['hanbin']

c1 = ws['A1']
c2 = ws['C9']

print (c1, c2)
