import openpyxl

xf = r'C:\Users\Administrator\AppData\Local\Temp\foo.xlsx'

wb = openpyxl.load_workbook(xf, data_only=True)
ws = wb['hanbin']
fc = ws['C54']
print('C54=%s' % fc.value)

import formulas
# fpath = 'file.xlsx'
xl_model = formulas.ExcelModel().loads(xf).finish()
xl_model.calculate()
xl_model.write()

wb = openpyxl.load_workbook(xf, data_only=True)
ws = wb['hanbin']
fc = ws['C54']
print('C54=%s' % fc.value)
#
# fc = ws['A54']
# print(fc.internal_value)
