import os
import sys
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
try:
    import win32com.client
    win32_ready = True
except Exception:
    win32_ready = False
import csv
import statistics as stats


################################################################################
class Excel:
    # ==========================================================================
    # 엑셀 파일 불러오기
    def load_file(self, filename):
        """
        :param filename: (str) 불러올 엑셀 파일 이름
        :return: <Workbook> 엑셀 파일의 워크북
        """
        if os.path.exists(filename):
            wb = load_workbook(filename, data_only=True)
        else:
            wb = Workbook()
        return wb

    # ==========================================================================
    # 워크시트 불러오기
    def load_worksheet(self, wb, sheetname):
        """
        :param wb: (Workbook) 워크시트를 부를 워크북
        :param sheetname: (str) 생성할 워크시트 이름
        :return: <Worksheet> 생성한 워크시트
        """
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
        elif sheetname == 'none':
            ws = wb.active
        else:
            ws = wb.create_sheet(sheetname)
        return ws

    # ==========================================================================
    # 데이터 읽어오기
    def get_value(self, ws, v, cell_range):
        """
        :param ws: (Worksheet) 데이터를 읽을 워크시트
        :param v: (bool) 데이터 읽을 방향(true이면 수직방향)
        :param cell_range: (str) 읽을 셀 범위
        :return: (list) 읽은 데이터의 리스트
        """
        val = []

        if cell_range is None:
            if v:
                for col in ws.columns:
                    tmp = []
                    for cell in col:
                        tmp.append(cell.value)
                    val.append(tmp)
            else:
                for row in ws.rows:
                    tmp = []
                    for cell in row:
                        tmp.append(cell.value)
                    val.append(tmp)
        # 단일 셀에 입력하는 경우
        elif ':' not in ''.join(map(str, cell_range)):
            val.append(ws[''.join(map(str, cell_range.split(':')))])
        else:
            for r in range(len(cell_range)):
                cell_range_tmp = ''.join(map(str, cell_range[r]))
                head_range = cell_range_tmp.split(':')[0]
                tail_range = cell_range_tmp.split(':')[1]
                min_col = ws[head_range].column
                # min_col = column_index_from_string(ws[head_range].column)
                min_row = ws[head_range].row
                # max_col = column_index_from_string(ws[tail_range].column)
                max_col = ws[tail_range].column
                max_row = ws[tail_range].row

                if v:
                    for row in ws.iter_cols(min_row=min_row, min_col=min_col,
                                            max_row=max_row, max_col=max_col):
                        tmp = []
                        for cell in row:
                            try:
                                # tmp.append(''.join(map(str, cell.value)))
                                tmp.append(cell.value)
                            except TypeError:
                                pass
                        val.append(tmp)
                else:
                    for col in ws.iter_rows(min_row=min_row, min_col=min_col,
                                            max_row=max_row, max_col=max_col):
                        tmp = []
                        for cell in col:
                            try:
                                tmp.append(cell.value)
                            except TypeError:
                                pass
                        val.append(tmp)
        return val

    # ==========================================================================
    # CSV로 변환
    def to_csv(self, data: list):
        """
        :param data: (list) csv로 변환할 값
        :return: True
        """
        wr = csv.writer(sys.stdout, lineterminator=os.linesep)
        for cell in data:
            if cell is None:
                # cell = ''
                continue
            wr.writerow(cell)

        return True

    # ==========================================================================
    # write mode
    # 엑셀 파일 로드-> 기존 파일 or 파일 생성 판별할 수 있어야함: load_file
    # 시트 로드 -> 기존 시트 or 시트 생성 판별할 수 있어야함: load_worksheet
    # 데이터 쓰기 & 엑셀 저장 -> 방향에 따라 구현하기
    def set_value(self, wb, ws, cell_range, filename, variable, v):
        """
        :param wb: (Workbook) 쓸 워크북 이름
        :param ws: (Worksheet) 쓸 워크시트 이름
        :param cell_range: (str) 쓸 셀의 범위
        :param filename: (str) 쓸 엑셀파일 이름
        :param variable: (list) 쓸 값
        :param v: (bool) 읽는 방향(true이면 수직방향)
        :return: (Workbook)
        """
        if cell_range is None:
            k = 0
            for row in ws.rows:
                for cell in row:
                    try:
                        cell.value = variable[k]
                        k += 1
                    except IndexError:
                        wb.save(filename)
                        return wb

        elif ':' not in ''.join(map(str, cell_range)):
            ws[''.join(map(str, cell_range.split(':')))] = variable[0]
        else:
            k = 0
            for r in range(len(cell_range)):
                cell_range_tmp = ''.join(map(str, cell_range[r]))
                head_range = cell_range_tmp.split(':')[0]
                tail_range = cell_range_tmp.split(':')[1]
                min_col = column_index_from_string(ws[head_range].column)
                min_row = ws[head_range].row
                max_col = column_index_from_string(ws[tail_range].column)
                max_row = ws[tail_range].row

                if v:
                    for i in range(min_col, max_col + 1):
                        for j in range(min_row, max_row + 1):
                            try:
                                ws.cell(row=j, column=i).value = variable[k]
                            except IndexError:
                                ws.cell(row=j, column=i).value = None
                            k += 1
                else:
                    for i in range(min_row, max_row + 1):
                        for j in range(min_col, max_col + 1):
                            try:
                                ws.cell(row=i, column=j).value = variable[k]
                            except IndexError:
                                ws.cell(row=i, column=j).value = None
                            k += 1
        wb.save(filename)
        return wb

    # ==========================================================================
    # function mode
    # csv file 열기
    def load_csv_file(self, filename):
        """
        :param filename: (str) 가져올 csv 파일 이름
        :return: (file) csv 파일
        """
        f = open(filename, 'r', encoding='utf-8')
        return f

    # ==========================================================================
    # csv 데이터 읽어오기
    def get_csv_value(self, file):
        """
        :param file: (file) 데이터를 읽을 csv 파일 이름
        :return: (list) csv 파일 내용
        """
        val = []

        for row in csv.reader(file):
            val.append(row)
        return val

    # ==========================================================================
    # excel로 변환 + 연산(function)
    def do_function(self, data: list, ftype, fargs, fcell):
        """
        :param data: (list) csv 파일에 있던 값
        :param ftype: (str) 값에 취할 연산 종류
        :param fargs: (str) 함수에 필요한 인자
        :param fcell: (str) 결과 값을 저장할 위치
        :return:
        """
        if os.path.exists('tmp.xlsx'):
            os.remove('tmp.xlsx')

        # 엑셀 파일 생성
        wb = self.load_file('tmp.xlsx')
        ws = self.load_worksheet(wb, 'none')

        # 받은 list가 str형이기 때문에 int(float)형으로 변환
        for i, val_i in enumerate(data):
            for j, val_j in enumerate(val_i):
                if self.is_number(data[i][j]) is True:
                    data[i][j] = float(data[i][j])

        # csv 데이터를 엑셀로 옮김
        for row in data:
            ws.append(row)

        # 연산식 삽입
        ws['A7'] = '={}({})'.format(ftype, fargs)
        ws[fcell] = ws['A7'].value

        wb.save('tmp.xlsx')

        wb = self.load_file('tmp.xlsx')
        ws = self.load_worksheet(wb, 'none')
        val = []
        for row in ws.rows:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            val.append(tmp)

        return val

    # ==========================================================================
    def get_function(self):
        # wb = self.load_file('tmp.xlsx')
        # ws = self.load_worksheet(wb, 'none')

        wb = load_workbook('tmp.xlsx', data_only=True)
        ws = wb.active

        val = []
        for row in ws.rows:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            val.append(tmp)

        os.remove('tmp.xlsx')
        return val

    # ==========================================================================
    def is_number(self, s):
        try:
            complex(s)  # for int, long, float and complex
        except ValueError:
            return False
        return True

    # ==========================================================================
    def do_macro(self, filename, macroname):
        """
        :param filename: (str) 매크로를 실행할 엑셀파일 이름
        :param macroname: (str) 실행할 매크로 이름
        :return: none
        """
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=os.path.abspath(filename), ReadOnly=1)
        xl.Application.Run(macroname)
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        del xl


################################################################################
def xlread(args):
    xl = Excel()

    wb = xl.load_file(args.filename)
    ws = xl.load_worksheet(wb=wb, sheetname=args.sheetname)
    data = xl.get_value(ws, args.vertical, args.readrange)
    xl.to_csv(data=data)


################################################################################
def xlwrite(args):
    xl = Excel()
    global lst
    lst = []

    for i, val_i in enumerate(args.input):
        lst.append(''.join(map(str, args.input[i])))
    print(type(lst))
    print(lst)

    wb = xl.load_file(args.filename)
    ws = xl.load_worksheet(wb=wb, sheetname=args.sheetname)
    xl.set_value(wb=wb, ws=ws, cell_range=args.writerange,
                 filename=args.filename, variable=lst, v=args.vertical)


################################################################################
def xlwritefunc(args):
    xl = Excel()

    f = xl.load_csv_file(args.filename)
    data = xl.get_csv_value(file=f)
    data = xl.do_function(data=data, ftype=args.func_name,
                          fargs=args.func_arg, fcell=args.func_location)
    data = xl.get_function()
    xl.to_csv(data=data)


################################################################################
def xlwritemacro(args):
    xl = Excel()
    xl.do_macro(args.filename, args.macroname)


################################################################################
def main():
    parser = argparse.ArgumentParser()

    parser.add_argument('filename', type=str, help='file name')
    parser.add_argument('--sheetname', type=str, help='sheet name',
                        default='none')
    parser.add_argument('-v', '--vertical', help='access vertically',
                        action='store_true')

    subparser = parser.add_subparsers()

    parser_r = subparser.add_parser('read')
    parser_r.add_argument('-r', '--readrange', type=str, help='range for read',
                          action='append')

    parser_w = subparser.add_parser('write')
    subparser_w = parser_w.add_subparsers()
    parser_f = subparser_w.add_parser('function')
    parser_f.add_argument('func_name', type=str, help='function name')
    parser_f.add_argument('func_arg', type=str, help='argumnets of function')
    parser_f.add_argument('func_location', type=str, help='result cell')

    parser_m = subparser_w.add_parser('macro')
    parser_m.add_argument('macroname', type=str, help='macro name')

    parser_i = subparser_w.add_parser('input')
    parser_i.add_argument('-w', '--writerange', type=str, help='range for write',
                          action='append')
    parser_i.add_argument('input', type=list, nargs='+', help='input value',
                          default=None)

    parser_r.set_defaults(func=xlread)
    parser_f.set_defaults(func=xlwritefunc)
    parser_m.set_defaults(func=xlwritemacro)
    parser_i.set_defaults(func=xlwrite)

    args = parser.parse_args()
    args.func(args)


################################################################################
if __name__ == "__main__":
    main()
