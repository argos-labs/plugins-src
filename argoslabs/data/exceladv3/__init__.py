#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.exceladv3'
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Formula
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
# * [2023/03/31] Kyobong
#     - fillfunc 함수에 m,cell 구하기위한 정규식에 패턴값이 숫자 한자리만 구할수 있음.
#       기존 "[a-zA-Z]+\d" -> 변경 "[a-zA-Z]+\d+"
# * [2021/10/08]
#     - newvalue base64 변환하는 부분이 에러가 발생해서 __init__ 부분에서 변환하도록 수정
# * [2021/03/26]
#     - update replace function
#  * [2020/08/10]
#     - build a plugin
#  * [2020/08/10]
#     - starting
#

################################################################################
import os
import re
import csv
import sys
import datetime
import openpyxl
import xlwings as xw
from tempfile import gettempdir
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit, vv_base64_decode


################################################################################
# noinspection PyBroadException
class Excel3API(object):
    OP_TYPE = ['Put value/formula', 'Replace value/formula', 'Convert str2num',
               'VLOOKUP', 'COUNT', 'COUNTA', 'COUNTIF', 'SUM', 'Fill formula',
               'Unmerge Cells']

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.newvalue = []
        if argspec.newvalue:
            for i in argspec.newvalue:
                self.newvalue.append(vv_base64_decode(i))
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.opened = False
        self.wb = None
        self.ws = None
        self.tempfile = None
        self.newcell = argspec.newcell

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    @staticmethod
    def datelst(cell, fmt_):
        DATETIME_FORMAT = {
            'YYYYMMDD-HHMMSS': "%Y%m%d-%H%M%S",
            'YYYY-MM-DD HH:MM:SS': "%Y-%m-%d %H:%M:%S",
            'YYYY/MM/DD HH:MM:SS': "%Y/%m/%d %H:%M:%S",
            'MMDDYYYY-HHMMSS': "%m%d%Y-%H%M%S",
            'MM-DD-YYYY HH:MM:SS': "%m-%d-%Y %H:%M:%S",
            'MM/DD/YYYY HH:MM:SS': "%m/%d/%Y %H:%M:%S",
            'M/D/YYYY HH:MM:SS': "%-m/%-d/%Y %H:%M:%S" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S",
            'YYYYMMDD-HHMMSS.mmm': "%Y%m%d-%H%M%S.%f",
            'YYYY-MM-DD HH:MM:SS.mmm': "%Y-%m-%d %H:%M:%S.%f",
            'YYYY/MM/DD HH:MM:SS.mmm': "%Y/%m/%d %H:%M:%S.%f",
            'MMDDYYYY-HHMMSS.mmm': "%m%d%Y-%H%M%S.%f",
            'MM-DD-YYYY HH:MM:SS.mmm': "%m-%d-%Y %H:%M:%S.%f",
            'MM/DD/YYYY HH:MM:SS.mmm': "%m/%d/%Y %H:%M:%S.%f",
            'M/D/YYYY HH:MM:SS.mmm': "%-m/%-d/%Y %H:%M:%S.%f" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S.%f",
            'YYYYMMDD': "%Y%m%d",
            'YYYY-MM-DD': "%Y-%m-%d",
            'YYYY/MM/DD': "%Y/%m/%d",
            'MMDDYYYY': "%m%d%Y",
            'MM-DD-YYYY': "%m-%d-%Y",
            'MM/DD/YYYY': "%m/%d/%Y",
            'M/D/YYYY': "%-m/%-d/%Y" if sys.platform != 'win32' else "%#m/%#d/%Y",
            'B D YYYY': "%b %-d %Y" if sys.platform != 'win32' else "%b %#d %Y",
            'B D, YYYY': "%b %-d, %Y" if sys.platform != 'win32' else "%b %#d, %Y",
            'D B YYYY': "%-d %b %Y" if sys.platform != 'win32' else "%#d %b %Y",
        }
        if '%' in fmt_:
            try:
                cell.value = float(cell.value.strip('%')) / 100
                cell.number_format = fmt_
            except Exception:
                pass
        else:

            try:
                p = DATETIME_FORMAT[fmt_]
                cell.value = datetime.datetime.strptime(cell.value, p)
                cell.number_format = fmt_
            except Exception:
                try:
                    cell.value = float(cell.value)
                except Exception:
                    pass
                cell.number_format = fmt_

    # ==========================================================================
    def xls2csv(self):
        if self.argspec.newfilename:
            self.filename = self.argspec.newfilename
        with open(self.filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
            for r in self.ws.rows:
                for i in r:
                    try:
                        self.ws[i.coordinate] = float(i.value)
                    except Exception:
                        pass

        else:
            self.wb = openpyxl.load_workbook(self.filename, read_only=read_only,
                                             data_only=False,
                                             keep_vba=keep_vba)
            if self.argspec.sheetname is None:
                self.ws = self.wb.active
            elif self.argspec.sheetname:
                self.ws = self.wb[self.argspec.sheetname]
        self.opened = True
        return self.opened

    # ==========================================================================
    def csv2excel(self, rng, fmt):
        rn = self.ws[rng]
        if ':' in rng:
            k = rng.split(':')
            k = [i.lower() for i in k]
            if not re.match("\w+\d", k[0]) and k[0] == k[1]:
                for i in rn:
                    self.datelst(i, fmt)
            else:
                for i in rn:
                    for j in i:
                        self.datelst(j, fmt)
        else:
            self.datelst(rn, fmt)
        return 0

    # ==========================================================================
    def save(self):
        if not self.argspec.newfilename:
            if self.extension == '.csv':
                self.wb.save(self.tempfile)
                self.xls2csv()
                self.close()
            else:
                self.wb.save(self.filename)

            return os.path.abspath(self.filename)
        else:
            if self.extension == '.csv':
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.wb.save(self.tempfile)
                    self.xls2csv()
                    self.close()
                else:
                    self.wb.save(self.argspec.newfilename)
            else:
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.xls2csv()
                else:
                    self.wb.save(self.argspec.newfilename)
                self.wb.close()
            return os.path.abspath(self.argspec.newfilename)

    # ==========================================================================
    def close(self):
        if self.tempfile and os.path.exists(self.tempfile):
            os.remove(self.tempfile)

    # ==========================================================================
    def calc_range(self, rng):
        if ':' in rng:
            k = rng.split(':')
            k = [i.lower() for i in k]
            if not re.match("\w+\d", k[0]) and k[0] == k[1]:
                return 'same_col'  # A:A
            else:
                return 'diff_col'
        else:
            if self.ws[rng]:
                return 'cell'
            else:
                raise RuntimeError('Cannot read the range %s' % rng)

    # ==========================================================================
    def paste(self, newvalue, range_=None, newcell=None):
        try:
            newvalue[0] = float(newvalue[0])
        except Exception:
            pass
        if newcell:
            self.ws[newcell] = newvalue[0]
        elif range_:
            for j in range_:
                k = self.calc_range(j)
                t = self.ws[j]
                if k == 'same_col':
                    for i in t:
                        self.ws[i.coordinate] = newvalue[0]
                elif k == 'diff_col':
                    for i in t:
                        for x in i:
                            self.ws[x.coordinate] = newvalue[0]
                else:
                    self.ws[j] = newvalue[0]
        print(self.save(), end='')

    # ==========================================================================
    def replace(self, range_=None):
        if not self.argspec.oldvalue:
            self.argspec.oldvalue = None
        if not self.newvalue:
            self.newvalue = ['']
        try:
            self.newvalue[0] = float(self.newvalue[0])
        except Exception:
            pass
        try:
            self.argspec.oldvalue = float(self.argspec.oldvalue)
        except Exception:
            pass
        if range_:
            for j in range_:
                k = self.calc_range(j)
                t = self.ws[j]
                if k == 'same_col':
                    for i in t:
                        if i.value == self.argspec.oldvalue:
                            i.value = self.newvalue[0]
                elif k == 'diff_col':
                    for i in t:
                        for x in i:
                            if x.value == self.argspec.oldvalue:
                                x.value = self.newvalue[0]
                else:
                    if self.ws[j].value == self.argspec.oldvalue:
                        self.ws[j].value = self.newvalue[0]
        else:
            for i in self.ws.iter_rows():
                for x in i:
                    if x.value == self.argspec.oldvalue:
                        x.value = self.newvalue[0]
                        # print(x.value)
        print(self.save(), end='')

    # GIVE A CONDITION Digits

    # ==========================================================================
    def convert(self, range_):  # formula will be changed to float
        for j in range_:
            k = self.calc_range(j)
            t = self.ws[j]
            if k == 'same_col':
                for i in t:
                    try:
                        self.ws[i.coordinate] = float(i.value)
                    except Exception:
                        pass
            elif k == 'diff_col':
                for i in t:
                    for x in i:
                        try:
                            self.ws[x.coordinate] = float(x.value)
                        except Exception:
                            pass
            else:
                try:
                    self.ws[j] = float(t.value)
                except Exception as err:
                    raise RuntimeError(err)
        print(self.save(), end='')

    # ==========================================================================
    def vlookup(self, cell, range_, ind, bool_):
        self.ws[self.newcell] = f'=VLOOKUP({cell},{range_[0]},{ind},{bool_})'

    # ==========================================================================
    def count(self, range_, newvalue=None):
        if newvalue:
            range_ += newvalue
        k = ','.join(range_)
        self.ws[self.newcell] = f'=COUNT({k})'

    # ==========================================================================
    def counta(self, range_):
        k = ','.join(range_)
        self.ws[self.newcell] = f'=COUNTA({k})'

    # ==========================================================================
    def countif(self, range_, condition):
        self.ws[self.newcell] = f'=COUNTIF({range_[0]},{condition})'

    # ==========================================================================
    def sumfunc(self, range_, newvalue=None):
        if newvalue:
            range_ += newvalue
        k = ','.join(range_)
        self.ws[self.newcell] = f'=SUM({k})'

    # ==========================================================================
    def colfunc(self, ln, cell, m):
        for i in ln:
            self.ws[i.coordinate] = m % tuple(cell)
            lst = []
            for j in cell:
                k = coordinate_from_string(j)
                newcol = get_column_letter(column_index_from_string(k[0]) + 1)
                t = [newcol, str(k[1])]
                lst.append(''.join(t))
            cell = re.findall("\w+\d", m % tuple(lst))

    # ==========================================================================
    def fillfunc(self, range_, newvalue):
        # 패턴 변경 "[a-zA-Z]+\d" -> "[a-zA-Z]+\d+"
        cell = re.findall("[a-zA-Z]+\d+", newvalue)
        m = re.sub("[a-zA-Z]+\d+", '%s', newvalue)
        ln = self.ws[range_]
        t0 = self.calc_range(range_)
        if t0 == 'same_col':
            lst = cell
            for i in ln:
                self.ws[i.coordinate] = m % tuple(lst)
                lst = []
                for j in cell:
                    k = coordinate_from_string(j)
                    t = [k[0], str(k[1] + 1)]
                    lst.append(''.join(t))
                cell = re.findall("[a-zA-Z]+\d+", m % tuple(lst))
        elif t0 == 'diff_col':
            for i in ln:
                self.colfunc(i, cell, m)
                lst = []
                for j in cell:
                    k = coordinate_from_string(j)
                    t = [k[0], str(k[1] + 1)]
                    lst.append(''.join(t))
                cell = re.findall("[a-zA-Z]+\d+", m % tuple(lst))
        else:
            self.ws[range_] = newvalue
        print(self.save(), end='')

    # ==========================================================================
    def xlwingsfunc(self):
        if self.extension == '.csv':
            self.wb.save(self.tempfile)
            self.save()
        else:
            self.wb.save(self.filename)
            self.save()
        app = xw.App(visible=False)
        try:
            if self.extension == '.csv':
                wbxl = app.books.open(self.tempfile)
            else:
                wbxl = app.books.open(self.filename)
        except Exception:
            raise RuntimeError('The excel file contains the unreadable content')
        if self.argspec.sheetname:
            ws1 = wbxl.sheets[self.argspec.sheetname]
        else:
            ws1 = wbxl.sheets[0]
        v = str(ws1.range(self.newcell).value)
        try:
            v = float(v)
            v_fp = v - int(v)
            if abs(v_fp) < 0.000001:
                v = int(v)
                print(v, end='')
        except Exception:
            print(v, end='')
        wbxl.close()
        self.close()

    # ==========================================================================
    def unmerge(self):
        if self.argspec.range:
            lst = self.argspec.range
        else:
            t = self.ws.merged_cells.ranges
            lst = []
            for i in t:
                lst.append(i)
        for i in lst:
            min_col, min_row, max_col, max_row = range_boundaries(str(i))
            top_left_cell_value = self.ws.cell(row=min_row,
                                               column=min_col).value

            self.ws.unmerge_cells(str(i))
            for row in self.ws.iter_rows(min_col=min_col, min_row=min_row,
                                         max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        if not self.argspec.newfilename:
            self.argspec.newfilename = self._get_safe_next_filename(self.filename)
        print(self.save(), end='')
        return 0

    # ==========================================================================
    def do(self, op):
        if not self.argspec.range:
            self.argspec.range = list()
        if op == self.OP_TYPE[0]:
            self.paste(self.newvalue, self.argspec.range,
                       self.argspec.newcell)
        elif op == self.OP_TYPE[1]:
            self.replace(self.argspec.range)
        elif op == self.OP_TYPE[2]:
            self.convert(self.argspec.range)
        elif op == self.OP_TYPE[3]:
            self.vlookup(self.argspec.targetcell, self.argspec.range,
                         self.argspec.index, self.argspec.bool)
            self.xlwingsfunc()
        elif op == self.OP_TYPE[4]:
            self.count(self.argspec.range, self.newvalue)
            self.xlwingsfunc()
        elif op == self.OP_TYPE[5]:
            self.counta(self.argspec.range)
            self.xlwingsfunc()
        elif op == self.OP_TYPE[6]:
            self.countif(self.argspec.range, self.argspec.condition)
            self.xlwingsfunc()
        elif op == self.OP_TYPE[7]:
            self.sumfunc(self.argspec.range, self.newvalue)
            self.xlwingsfunc()
        elif op == self.OP_TYPE[8]:
            r, n = self.argspec.range, self.newvalue
            if len(r)!=len(n):
                raise RuntimeError('The length of Cell/Range and New value are different.')
            else:
                for i in range(len(r)):
                    # v = base64.b64decode(n[i]).decode('utf-8')
                    v = vv_base64_decode(n[i])
                    self.fillfunc(r[i], v)
        elif op == self.OP_TYPE[9]:
            self.unmerge()


################################################################################
@func_log
def do_excel3(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        exl = Excel3API(argspec)
        if exl.extension in ('.xlsx', '.csv'):
            exl.open()
        elif exl.extension == '.xlsm':
            exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if argspec.format_range and argspec.format:
            for i, j in zip(argspec.format_range, argspec.format):
                exl.csv2excel(i, j)
        exl.do(argspec.op)
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        # print(False)
        return 1
    finally:
        if exl is not None:
            exl.close()
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################

def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel AdvIII',
            icon_path=get_icon_path(__file__),
            description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ######################################### for app dependent options
        mcxt.add_argument('--sheetname', display_name='Sheetname',
                          help='Sheet name to handle. If not specified last'
                               ' activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          input_method='fileread', help='New filename')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newcell', display_name='Cell for formula',
                          help='New cell to save excel formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--oldvalue', display_name='Replace old Value',
                          help='Oldvalue to replace')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newvalue', display_name='New value',
                          action='append', input_method='base64',
                          help='New value to replace')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', display_name='Cell/Range', action='append',
                          help='Cell or range in excel file')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--format_range', display_name='Format Cell/Range',
                          action='append',
                          help='Cell or range change format')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--format', display_name='Format', action='append',
                          help='Specify cell value format e.g. YYYY-MM-DD')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--targetcell', display_name='VLOOKUP Target Cell',
                          help='Target cell in VLOOKUP')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--index', display_name='VLOOKUP Index',
                          help='Index in VLOOKUP')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--bool', display_name='VLOOKUP True', default=False,
                          type=bool,
                          help='TRUE/FALSE in VLOOKUP')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--condition', display_name='COUNIF Condition',
                          help='Condition for COUNTIF function')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')

        # ###################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big'
                               ' CSV input can take time.)')
        # ----------------------------------------------------------------------
        mcxt.add_argument('op',
                          display_name='Excel function type',
                          choices=Excel3API.OP_TYPE,
                          help='Excel advanced2 type of operation')
        argspec = mcxt.parse_args(args)
        return do_excel3(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
