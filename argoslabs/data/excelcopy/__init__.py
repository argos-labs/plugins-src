#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excelcopy'
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Copy Format
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2020/09/14]
#     - build a plugin
#  * [2020/09/14]
#     - starting
#

################################################################################
import os
import re
import csv
import sys
import shutil
import openpyxl
from copy import copy
from tempfile import gettempdir
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit


################################################################################
# noinspection PyBroadException
class ExcelcopyAPI(object):

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.opened = False
        self.wb = None
        self.ws = None
        self.ws0 = None
        self.tempfile = None
        self.range = self.argspec.range
        self.pasterange = self.argspec.pasterange
        self.cellval = self.argspec.copyval
        # self.insertrow = self.argspec.insertrow
        # self.insertcol = self.argspec.insertcol

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    @staticmethod
    def excelformat(vi, ti, val=None):
        vi.font = copy(ti.font)
        vi._style = copy(ti._style)
        vi.border = copy(ti.border)
        vi.number_format = copy(ti.number_format)
        vi.protection = copy(ti.protection)
        vi.alignment = copy(ti.alignment)
        if val:
            vi.value = ti.value
        return vi, ti

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def xls2csv(self):
        if self.argspec.newfilename:
            self.filename = self.argspec.newfilename
        with open(self.filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.argspec.dataonly:
            data_only = True
            if not self.argspec.newfilename:
                t = self._get_safe_next_filename(self.filename)
                shutil.copy(self.filename, t)
                self.filename = t
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
            for r in self.ws.rows:
                for i in r:
                    try:
                        self.ws[i.coordinate] = float(i.value)
                    except Exception:
                        pass

        else:
            self.wb = openpyxl.load_workbook(self.filename, read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if self.argspec.sheetname is None:
                self.ws = self.wb.active
            elif self.argspec.sheetname:
                self.ws = self.wb[self.argspec.sheetname]
            if self.argspec.pastesheetname:
                try:
                    self.ws0 = self.wb[self.argspec.pastesheetname]
                except Exception:
                    self.wb.create_sheet(self.argspec.pastesheetname)
                    self.ws0 = self.wb[self.argspec.pastesheetname]
        # if self.insertrow:
        #     try:
        #         self.ws.move_range(self.insertrow,
        #                            rows=len(self.ws[self.insertrow]))
        #     except:
        #         self.ws.move_range(self.insertrow, rows=1)
        #     self.pasterange = self.insertrow
        # elif self.insertcol:
        #     try:
        #         self.ws.move_range(self.insertcol, cols=len(self.ws[self.insertcol]))
        #     except:
        #         self.ws.move_range(self.insertcol, cols=1)
        #     self.pasterange = self.insertcol
        self.opened = True
        return self.opened

    # ==========================================================================
    def save(self):
        if not self.argspec.newfilename:
            if self.extension == '.csv':
                self.wb.save(self.tempfile)
                self.xls2csv()
                self.close()
            else:
                self.wb.save(self.filename)
            return os.path.abspath(self.filename)
        else:
            if self.extension == '.csv':
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.wb.save(self.tempfile)
                    self.xls2csv()
                    self.close()
                else:
                    self.wb.save(self.argspec.newfilename)
            else:
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.xls2csv()
                else:
                    self.wb.save(self.argspec.newfilename)
                self.wb.close()
            return os.path.abspath(self.argspec.newfilename)

    # ==========================================================================
    def close(self):
        if self.tempfile and os.path.exists(self.tempfile):
            os.remove(self.tempfile)

    # ==========================================================================
    def calc_range(self, rng):
        if ':' in rng:
            k = rng.split(':')
            k = [i.lower() for i in k]
            if not re.match("\w+\d", k[0]) and k[0] == k[1]:
                return 'same_col'
            else:
                return 'diff_col'
        else:
            if self.ws[rng]:
                return 'cell'
            else:
                raise RuntimeError('Cannot read the range %s' % rng)

    # ==========================================================================
    def copyformat(self):
        if self.range and self.pasterange:
            t = self.ws[self.range]
            if self.argspec.pastesheetname:
                v = self.ws0[self.pasterange]
            else:
                v = self.ws[self.pasterange]
            t_rng, v_rng, = self.calc_range(self.range), self.calc_range(
                self.pasterange)
            if v_rng == 'cell':
                val, minlen = v, 1
            elif t_rng == 'same_col' and v_rng == 'diff_col':
                t = [t] * len(v)
                t_rng = 'diff_col'
                val, minlen = v, len(v)
            elif t_rng == 'diff_col' and v_rng == 'diff_col':
                lt, lv, lt0, lv0 = len(t), len(v), len(t[0]), len(v[0])
                p, q = [], []
                if lt0 < lv0:
                    for i in range(len(t)):
                        if lv0 % lt0 == 0:
                            p.append(t[i] * int(lv0 / lt0))
                        else:
                            p.append(t[i] * (int(lv0 / lt0) + 1))
                    t = p * lt
                if lt < lv:
                    if lv % lt == 0:
                        q.append(t * int(lv / lt))
                    else:
                        q.append(t * (int(lv / lt) + 1))
                    t = q[0]
                val, minlen = v, len(v)
            else:
                val, minlen = v, len(v)
            for i in range(minlen):
                if t_rng == 'same_col' and v_rng == 'same_col':
                    vi, ti = v[i], t[i]
                    self.excelformat(vi, ti, self.cellval)
                elif t_rng == 'diff_col' and v_rng == 'same_col':
                    vi, ti = v[i], t[0][i]
                    self.excelformat(vi, ti, self.cellval)
                elif t_rng == 'diff_col' and v_rng == 'diff_col':
                    for j in range(len(val[i])):
                        vi, ti = v[i][j], t[i][j]
                        self.excelformat(vi, ti, self.cellval)
                elif t_rng == 'cell' and v_rng != 'cell':
                    for j in range(len(val[i])):
                        vi, ti = v[i][j], t
                        self.excelformat(vi, ti, self.cellval)
                elif t_rng == 'cell' and v_rng == 'cell':
                    v._style = copy(t._style)
                    if self.cellval:
                        v.value = t.value
                else:
                    if t_rng == 'diff_col':
                        v._style = copy(t[0][0]._style)
                        if self.cellval:
                            v.value = t[0][0].value
                    else:
                        v._style = copy(t[0]._style)
                        if self.cellval:
                            v.value = t[0].value

    # ==========================================================================
    def printvalue(self):
        if self.argspec.pastesheetname:
            self.ws0.print_area = self.argspec.printarea
        else:
            self.ws.print_area = self.argspec.printarea


################################################################################
@func_log
def do_excelcopy(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        exl = ExcelcopyAPI(argspec)
        if exl.extension in ('.xlsx', '.csv'):
            exl.open()
        elif exl.extension == '.xlsm':
            exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if argspec.pasterange and argspec.range:
            exl.copyformat()
        # if argspec.copyvalue or argspec.insertrow or argspec.insertcol:
        # if argspec.copyvalue:
        #     exl.paste()
        if argspec.printarea:
            exl.printvalue()
        print(exl.save(), end='')
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        # print(False)
        return 1
    finally:
        if exl is not None:
            exl.close()
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################

def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Copy Paste',
            icon_path=get_icon_path(__file__),
            description='Excel copy format and select print area',
    ) as mcxt:
        # ######################################### for app dependent options
        mcxt.add_argument('--sheetname', display_name='Copy from Sheet',
                          help='Sheet name to handle. If not specified last'
                               ' activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--pastesheetname', display_name='Paste to Sheet',
                          help='Sheet name to paste')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='newfilename')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', display_name='Copy from Cell/Range',
                          help='range of the excel file ')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--pasterange', display_name='Paste to Cell/Range',
                          help='range of the excel file to paste')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--copyval', display_name='Copy with values',
                          action='store_true',
                          help='copy the value of cells')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--dataonly', display_name='Data Only',
                          action='store_true',
                          help='print only data not formulas')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--printarea', display_name='Set Print Area',
                          help='select print area in excel')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')

        # ###################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big'
                               ' CSV input can take time.)')
        argspec = mcxt.parse_args(args)
        return do_excelcopy(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
