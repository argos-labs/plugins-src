#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.exceladv2`
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Advance II
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2023/03/02] Kyobong An
#     - def insert_del_row_col 수정
#       - insert row/col and delete row/col 옵션에 'A-C,E,G'처럼 여러행 처리할수있도록 변경
#       - 삭제하면 왼쪽이나 위로 정렬됨 따라서 역순으로 제거해줘야함.
#  * [2023/01/27] Kyobong An
#     - 이전에 수정한 Sheet Copy에 문제점이 있었음. 수전전 코드 다시 복구
#     - data_only시 새로운 workbook을 불러서 시트만 복사하는 코드 추가
#       기존 수식이 날라가는 현상은 부르기전에 xlwing를 사용해 한번 저장해준후 진행
#  * [2021/11/10] Kyobong An
#     - Sheet Copy 할때 기존시트가 같이 생성되고 data_only까지 사용할때 기존파일이 여러개생기는 오류수정.
#  * [2021/03/29]
#     - 그룹에 "2-Business Apps" 넣음
#  * [2020/07/27]
#     - update the conversion error between csv and xlsx
#  * [2020/04/13]
#     - build a plugin
#  * [2020/04/06]
#     - starting
#

################################################################################
import os
import csv
import sys
import shutil
import win32com.client
from tempfile import gettempdir
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit
# noinspection PyPackageRequirements
import openpyxl
import xlwings
# noinspection PyPackageRequirements
from openpyxl.utils.cell import column_index_from_string, \
    coordinate_from_string, get_column_letter
from io import StringIO

################################################################################
ENCODINGS = [
    "auto: Detecting",
    "cp932: Japanese",
    "eucjp: Japanese",
    "cp936: Chinese (simplified)",
    "cp949: Korean",
    "euckr: Korean",
    "cp950: Chinese (traditional)",
    "utf-8 UTF-8",
]


################################################################################
class Excel2API(object):
    OP_TYPE = [
        'Save As', 'Add sheet', 'Copy Sheet', 'Rename sheet', 'Delete sheet',
        'Print PDF',
        'Insert delete row col', 'Find first blank col', 'Find first blank row'
    ]

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.newfilename = self.argspec.newfilename
        self.wb = None
        self.ws = None
        self.tempfile = None
        self.rr = None

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def __repr__(self):
        sio = StringIO()
        wr = csv.writer(sio, lineterminator='\n')
        for row in self.rr:
            wr.writerow(row)
        return sio.getvalue()

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    @staticmethod
    def col_name(colname):
        amount = 0
        lst = colname.strip().split('-')
        if len(lst) > 1:
            k = []
            for i in lst:
                k.append(column_index_from_string(i))
            amount = k[1] - k[0] + 1
            colname = k[0]
        elif len(lst) == 1:
            colname = column_index_from_string(colname)
            amount = 1
        return colname, amount

    # ==========================================================================
    @staticmethod
    def row_num(rownum):
        amount = 0
        lst = rownum.strip().split('-')
        if len(lst) > 1:
            amount = int(lst[1]) - int(lst[0]) + 1
            rownum = int(lst[0])
        elif len(lst) == 1:
            rownum = int(rownum)
            amount = 1
        return rownum, amount

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.argspec.data_only and self.argspec.op != 'Copy Sheet':
            t = self._get_safe_next_filename(self.filename)
            shutil.copy(self.filename, t)
            self.filename = t
            data_only = True

        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.load_workbook(self.filename,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if self.argspec.sheetname is None:
                self.ws = self.wb.active
            elif self.argspec.sheetname:
                if self.argspec.sheetname not in self.wb.sheetnames:
                    return 1
                self.ws = self.wb[self.argspec.sheetname]

        return 0

    # ==========================================================================
    def close(self):
        if self.tempfile and os.path.exists(self.tempfile):
            os.remove(self.tempfile)

    # ==========================================================================
    def xls2csv(self):
        if self.newfilename:
            self.filename = self.newfilename
        with open(self.filename, 'w', encoding=self.argspec.encoding) as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
            self.wb.close()
        return True

    # ==========================================================================
    def save(self):
        if not self.newfilename:
            if self.extension == '.csv':
                self.wb.save(self.tempfile)
                self.xls2csv()
                self.close()
            else:
                self.wb.save(self.filename)
            print(os.path.abspath(self.filename), end='')
            # return os.path.abspath(self.filename)
        else:
            if self.argspec.data_only and self.argspec.op != 'Copy Sheet':
                if not os.path.exists(self.newfilename):
                    os.remove(self.filename)
                else:
                    self.newfilename = self.filename
            if self.extension == '.csv':
                if self.newfilename.lower().endswith('csv'):
                    self.wb.save(self.tempfile)
                    self.xls2csv()
                    self.close()
                else:
                    self.wb.save(self.newfilename)
            else:
                if self.newfilename.lower().endswith('csv'):
                    self.xls2csv()
                else:
                    self.wb.save(self.newfilename)
                self.wb.close()
            print(os.path.abspath(self.newfilename), end='')
            # return os.path.abspath(self.newfilename)

    # ==========================================================================
    def addsheet(self):
        self.wb.create_sheet(self.argspec.newsheet)
        self.save()

    # ==========================================================================
    def rename_sheet(self):
        self.ws.title = self.argspec.newsheet
        self.save()

    # ==========================================================================
    def delsheet(self):
        self.wb.remove(self.ws)
        self.save()

    # ==========================================================================
    def printpdf(self):
        path_to_pdf = self.argspec.filenamepath
        wn = win32com.client.Dispatch('Excel.Application')
        wn.Visible = False
        if self.extension == 'csv':
            filename = self.tempfile
            self.wb.close()
        else:
            filename = self.filename
        wb = wn.Workbooks.Open(filename)
        if self.argspec.sheetname:
            wb.Worksheets(self.argspec.sheetname).Activate()
            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
        else:
            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
        wb.Close(True)
        print(os.path.abspath(path_to_pdf), end='')

    # ==========================================================================
    def insert_del_row_col(self):
        if self.argspec.col_name:
            col_names = self.argspec.col_name.split(',')
            for col_name in col_names:
                colname, amount = self.col_name(col_name)
                self.ws.insert_cols(colname, amount)
        if self.argspec.row_num:
            row_nums = self.argspec.row_num.split(',')
            for row_num in row_nums:
                rownum, amount = self.row_num(row_num)
                self.ws.insert_rows(rownum, amount)
        if self.argspec.del_col_name:
            del_col_names = self.argspec.del_col_name.split(',')
            for del_col_name in reversed(del_col_names):
                colname, amount = self.col_name(del_col_name)
                self.ws.delete_cols(colname, amount)
        if self.argspec.del_row_num:
            del_row_nums = self.argspec.del_row_num.split(',')
            for del_row_num in reversed(del_row_nums):
                rownum, amount = self.row_num(del_row_num)
                self.ws.delete_rows(rownum, amount)
        self.save()

    # ==========================================================================
    def blank_col(self):
        result = ''
        for i in self.ws[int(self.argspec.row_num)]:
            if i.value is None:
                if self.argspec.nonecsvrv:
                    result = get_column_letter(i.column)+str(i.row)
                else:
                    result = f'{get_column_letter(i.column)},{i.row}'
                break
            elif self.ws.max_column + 1:
                col = self.ws.max_column + 1
                if self.argspec.nonecsvrv:
                    result = get_column_letter(col)+str(int(self.argspec.row_num))
                else:
                    result = f'{get_column_letter(col)},{int(self.argspec.row_num)}'
            else:
                result = "No blank was found"
        return result

    # ==========================================================================
    def blank_row(self):
        result = ''
        for i in self.ws[self.argspec.col_name]:
            if i.value is None:
                xy = coordinate_from_string(i.coordinate)
                if self.argspec.nonecsvrv:
                    result = str(xy[0])+str(xy[1])
                else:
                    result = f'{xy[0]},{xy[1]}'
                break
            elif self.ws.max_row + 1:
                if self.argspec.nonecsvrv:
                    result = self.argspec.col_name+str(self.ws.max_row + 1)
                else:
                    result = f'{self.argspec.col_name},{self.ws.max_row + 1}'
            else:
                result = "No blank was found"
        return result

    # ==========================================================================
    def copy_sheet(self):
        self.wbt = None
        if self.argspec.data_only:
            excel_app = xlwings.App(visible=False)
            excel_book = excel_app.books.open(self.filename)
            excel_book.save(self.filename)
            excel_book.close()
            excel_app.quit()
            self.wbt = openpyxl.load_workbook(self.filename, data_only=True)
            target = self.wbt[self.argspec.sheetname]
            target.title = 'copy' + target.title
            target._parent = self.wb
            self.wb._add_sheet(target)
            self.ws = self.wb[target.title]
        else:
            target = self.wb[self.argspec.sheetname]
            self.ws = self.wb.copy_worksheet(target)
        if self.argspec.newsheet:
            self.ws.title = self.argspec.newsheet
        self.save()
        if self.wbt:
            self.wbt.close()
        #self.wb.save('sample0.xlsx')

    # ==========================================================================
    def do(self, op):
        if op == 'Save As':
            self.save()
        if op == 'Add sheet':
            self.addsheet()
        if op == 'Rename sheet':
            self.rename_sheet()
        if op == 'Delete sheet':
            self.delsheet()
        if op == 'Print PDF':
            self.printpdf()
        if op == 'Insert delete row col':
            self.insert_del_row_col()
        if op == 'Find first blank col':
            print(self.blank_col(), end='')
        if op == 'Find first blank row':
            print(self.blank_row(), end='')
        if op == 'Copy Sheet':
            self.copy_sheet()

################################################################################
@func_log
def do_excek2(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        exl = Excel2API(argspec)
        if exl.extension in ('.xlsx', '.csv'):
            exl.open()
        elif exl.extension == '.xlsm':
            exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if exl.open()==1:
            print(f'Cannot find the sheetname "{argspec.sheetname}"', end='')
            return exl.open()
        exl.do(argspec.op)
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        # print(False)
        return 9
    finally:
        if exl is not None:
            exl.close()
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',  # Business Apps
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel AdvII',
            icon_path=get_icon_path(__file__),
            description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ######################################## for app dependent options
        mcxt.add_argument('--sheetname', display_name='Sheetname',
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='newfilename')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newsheet', display_name='New sheetname',
                          help='newsheet')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--filenamepath', display_name='PDF filename',
                          help='filenamepath', input_method='filewrite')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--col-name', display_name='Colname',
                          help='insert column name')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--row-num', display_name='Rowid',
                          help='insert row id')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--del-col-name', display_name='Del colname',
                          help='del column name')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--del-row-num', display_name='Del rowid',
                          help='del row num')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding', default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--data-only', display_name='Data Only',
                          action='store_true', help='data_only')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--nonecsvrv', display_name='None CSV RV',
                          action='store_true', help='print cell')
        # ##################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')
        # ----------------------------------------------------------------------
        mcxt.add_argument('op',
                          display_name='Excel function type',
                          choices=Excel2API.OP_TYPE,
                          help='Excel advanced2 type of operation')
        argspec = mcxt.parse_args(args)
        return do_excek2(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
