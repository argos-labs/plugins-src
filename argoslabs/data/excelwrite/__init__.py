#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excelwrite`
======0==============================
.. moduleauthor:: Kyobong An <akb0930@argos-labs.com>, Irene Cho <irene@argos-labs.com> 0
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel
"""
#
# Authors
# ===========
#
# * Kyobong An, Irene Cho
#
# Change Log
# --------
#  * [2023/03/16]
#   data only 사용시 xlwing에서 간헐적으로 sheet를 찾기 못하는 에러발생. sheet를 찾는 로직 추가
#   "Sheet Finder" 기능 추가 기본값은 5
#  * [2023/03/16]
#   "Add value" 기능 추가
#  * [2021/05/26]
#   input이 excelvalue로 들어오고 output을 csv로 나갈때 오류나던 부분, csv값이 사라지던 부분 모두 수정완료
#  * [2021/05/13]
#   엑셀이 닫히지 않는 오류 수정
#  * [2021/05/13]
#   csv-> xlsx, csv -> csv, xlsx -> csv, 확장자 변경하는 부분 수정완료
#  * [2021/05/12]
#   refactoring, sheet 추가하는 부분 오류수정, excel_value에서 범위 설정하는 부분 수정
#  * [2021/04/26]
#   데이터 타입 비교부분 수정
#  * [2021/04/16]
#   excel_value= 기능추가, blank도 들어갈 수 있음, data type 설정
#  * [2021/04/14]
#   xlwings 추가
#  * [2021/04/05]
#   실수타입을 텍스트형식으로 저장하는 것 수정
#  * [2021/04/05]
#   return value -> output 경로로 변경
#  * [2021/04/02]
#   수정요청사항 input/ output 라벨 변경, value 삭제
#  * [2021/04/01]
#   data only 기능추가, 숫자형을 텍스트형식으로 저장하는 것 수정
#  * [2021/03/19]
#   단일행 오류 수정
#  * [2021/03/18]
#   행과열 전체 읽기 추가


################################################################################
import os
import sys
import csv
# import time
import datetime
import time

# import shutil
import openpyxl
import xlwings
from openpyxl import Workbook
from tempfile import gettempdir
from openpyxl.utils import get_column_letter
from io import StringIO
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import ModuleContext, func_log, \
    ArgsError, ArgsExit, get_icon_path
import warnings

warnings.filterwarnings("ignore")

DATETIME_FORMAT = {
    'YYYYMMDD-HHMMSS': "%Y%m%d-%H%M%S",
    'YYYY-MM-DD HH:MM:SS': "%Y-%m-%d %H:%M:%S",
    'YYYY/MM/DD HH:MM:SS': "%Y/%m/%d %H:%M:%S",
    'MMDDYYYY-HHMMSS': "%m%d%Y-%H%M%S",
    'MM-DD-YYYY HH:MM:SS': "%m-%d-%Y %H:%M:%S",
    'MM/DD/YYYY HH:MM:SS': "%m/%d/%Y %H:%M:%S",
    'M/D/YYYY HH:MM:SS': "%-m/%-d/%Y %H:%M:%S" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S",
    'YYYYMMDD-HHMMSS.mmm': "%Y%m%d-%H%M%S.%f",
    'YYYY-MM-DD HH:MM:SS.mmm': "%Y-%m-%d %H:%M:%S.%f",
    'YYYY/MM/DD HH:MM:SS.mmm': "%Y/%m/%d %H:%M:%S.%f",
    'MMDDYYYY-HHMMSS.mmm': "%m%d%Y-%H%M%S.%f",
    'MM-DD-YYYY HH:MM:SS.mmm': "%m-%d-%Y %H:%M:%S.%f",
    'MM/DD/YYYY HH:MM:SS.mmm': "%m/%d/%Y %H:%M:%S.%f",
    'M/D/YYYY HH:MM:SS.mmm': "%-m/%-d/%Y %H:%M:%S.%f" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S.%f",
    'YYYYMMDD': "%Y%m%d",
    'YYYY-MM-DD': "%Y-%m-%d",
    'YYYY/MM/DD': "%Y/%m/%d",
    'MMDDYYYY': "%m%d%Y",
    'MM-DD-YYYY': "%m-%d-%Y",
    'MM/DD/YYYY': "%m/%d/%Y",
    'M/D/YYYY': "%-m/%-d/%Y" if sys.platform != 'win32' else "%#m/%#d/%Y",
    'B D YYYY': "%b %-d %Y" if sys.platform != 'win32' else "%b %#d %Y",
    'B D, YYYY': "%b %-d, %Y" if sys.platform != 'win32' else "%b %#d, %Y",
    'D B YYYY': "%-d %b %Y" if sys.platform != 'win32' else "%#d %b %Y",
}


################################################################################
class OpenError(Exception):
    pass


################################################################################
class Excel(object):
    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        self.outfile = argspec.write
        if not argspec.excel_value and not os.path.exists(self.filename):
            raise IOError('Cannot read excel file "%s"' % self.filename)
        if argspec.excel_value:
            self.extension = '.xlsx'
        else:
            self.extension = self._get_extension(self.filename)
        self.wb = None
        self.ws = None
        self.o_wb = None
        self.o_ws = None
        self.min_row = self.max_row = -1
        self.min_col = self.max_col = -1
        self.o_min_row = self.o_max_row = -1
        self.o_min_col = self.o_max_col = -1
        self.rr = None
        self.rrt = None
        # for csv reading
        self.tempfile = None
        self.temp_cell = ''

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def __repr__(self):
        sio = StringIO()
        wr = csv.writer(sio, lineterminator='\n')
        for row in self.rr:
            wr.writerow(row)
        return sio.getvalue()

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, 'rt', encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
            f.close()
            wb.close()
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def xls2csv(self):
        with open(self.outfile, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.o_ws.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.filename:
            self.extension = self._get_extension(self.filename)
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
        else:
            if self.filename:
                self.wb = openpyxl.load_workbook(self.filename,
                                                 read_only=read_only,
                                                 data_only=data_only,
                                                 keep_vba=keep_vba)
            else:
                self.wb = openpyxl.Workbook()
            if not self.argspec.sheet:
                self.ws = self.wb.active
            elif self.argspec.sheet in self.wb.sheetnames:
                self.ws = self.wb[self.argspec.sheet]
            else:
                raise OpenError(
                    f'The sheet name "{self.argspec.sheetname}" does not exist')
        return True

    # ==========================================================================
    def open_outputfile(self, read_only=False, data_only=False, keep_vba=False):
        self.extension = self._get_extension(self.argspec.write)
        if self.extension == '.csv' and os.path.exists(self.outfile):
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.outfile)[:-4])
            self.csv2xls(self.outfile, self.tempfile, self.argspec.encoding)
            self.o_wb = openpyxl.load_workbook(self.tempfile,
                                               read_only=read_only,
                                               data_only=data_only,
                                               keep_vba=keep_vba)
            if self.argspec.write_sheet:
                self.o_ws = self.o_wb.create_sheet()
                self.o_ws.title = self.argspec.write_sheet
            else:
                self.o_ws = self.o_wb.active
        else:
            if os.path.exists(self.outfile):
                self.o_wb = openpyxl.load_workbook(self.outfile,
                                                   read_only=read_only,
                                                   data_only=data_only,
                                                   keep_vba=keep_vba)
            else:
                self.o_wb = openpyxl.Workbook()
            if not self.argspec.write_sheet:
                self.o_ws = self.o_wb.active
            elif self.argspec.write_sheet in self.o_wb.sheetnames:
                self.o_ws = self.o_wb[self.argspec.write_sheet]
            else:
                self.o_ws = self.o_wb.create_sheet()
                self.o_ws.title = self.argspec.write_sheet
        if self.argspec.excel_value:
            self.ws = self.o_ws

        return True

    # ==========================================================================
    def close(self):
        if self.open:
            self.wb.close()
            if self.tempfile and os.path.exists(self.tempfile):
                os.remove(self.tempfile)
            self.ws = None
            self.wb = None

            return True
        return False

    # ==========================================================================
    def range_len1(self):  # 단일 행이나 열을 체크 ex) 'A'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha():
            self.rr[0] = self.rr[0] + '1'
        elif self.rr[0].isdigit():
            self.rr[0] = 'A' + self.rr[0]
        else:
            self.rr = self.rr * 2

    def range_len3(self):  # 두줄이상의 행이나 열체크  ex) 'A:B','2:16'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha() and self.rr[1].isalpha():
            self.rr[0] = self.rr[0] + '1'
            self.rr[1] = self.rr[1] + '1'
        elif self.rr[0].isdigit() and self.rr[1].isdigit():
            self.rr[0] = 'A' + self.rr[0]
            self.rr[1] = 'A' + self.rr[1]
        else:
            pass

    # ==========================================================================
    def calc_range(self):
        if not self.open:
            return False
        # 범위가 None 값일때 전체범위
        self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
        self.min_col, self.max_col = self.ws.min_column, self.ws.max_column
        self.o_min_row, self.o_max_row = self.o_ws.min_row, self.o_ws.max_row
        self.o_min_col, self.o_max_col = self.o_ws.min_column, self.o_ws.max_column
        if not self.argspec.range:
            self.temp_cell = 'A1'
            return False
        rrt = self.argspec.range.strip().split(':')
        self.temp_cell = self.rr[0]
        if self.argspec.range.isalnum():  # 단일행이나 열의 볌위 'A'
            if self.argspec.range.isalpha():
                c = self.ws[self.rr[0]]
                self.min_col, self.max_col = c.column, c.column
                return True
            elif self.argspec.range.isdigit():
                c = self.ws[self.rr[0]]
                self.min_row, self.max_row = c.row, c.row
                return True

        if rrt[0].isalpha() and rrt[1].isalpha():  # 범위  'A:B'  처럼 열범위
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_col, self.max_col = min(c1.column, c2.column), max(
                c1.column, c2.column)

        elif rrt[0].isdigit() and rrt[1].isdigit():  # 범위 '1:16' 처럼 행범위체크
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row,
                                                                  c2.row)
        else:  # 특정범위를 지정  'A4:E4'
            if self.argspec.range.isalnum():
                c1, c2 = self.ws[self.rr[0].strip()], self.ws[
                    self.rr[1].strip()]
            else:
                r1, r2 = self.argspec.range.strip().split(':')
                c1, c2 = self.ws[r1.strip()], self.ws[r2.strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row,
                                                                  c2.row)
            self.min_col, self.max_col = min(c1.column, c2.column), max(
                c1.column, c2.column)
        return True

    # ==========================================================================
    def save(self):
        argspec_write = self.outfile
        if self.extension == '.csv':  # 만약 결과를 csv에 쓴다면
            if os.path.exists(self.outfile):
                self.o_wb.save(self.outfile)
            else:
                self.o_wb.save(self.outfile)
            self.xls2csv()
        else:  # 나머지는 모두 excel 파일이라 생각
            self.o_wb.save(argspec_write)
        self.wb.close()
        self.o_wb.close()
        self.o_ws = None
        self.o_wb = None
        print(os.path.abspath(argspec_write), end='')
        return 0

    # ==========================================================================
    @staticmethod
    def value_type(value, fmt_=None):
        if not value:
            return value
        elif isinstance(value, datetime.datetime):
            try:
                p = DATETIME_FORMAT[fmt_]
                value = value.strftime(p)
            except Exception:
                pass
        else:
            try:
                value = float(value)
            except Exception as e:
                t = str(e).split(':')[1].strip()
                if "ufeff3" in t:
                    raise OpenError(f'Cannot read the value {t}. Check the encoding of the file.')
                pass
        return value

    # ==========================================================================
    def overwrite_cell(self, mcxt):
        try:
            ws1, wbxl = None, None
            if self.argspec.data_only:
                app = xlwings.App(visible=False)
                wbxl = app.books.open(self.filename)
                if self.argspec.sheet:
                    for i in range(self.argspec.sheetfind):
                        try:
                            ws1 = wbxl.sheets[self.argspec.sheet]
                            # sheet를 정상적으로 불러왔는지 확인. 간헐적으로 가져오지못함.
                            if ws1.range('A1').value:
                                break
                        except Exception as err:
                            # print(f'retry {i}')
                            mcxt.logger.error(f'get sheet retry {i}')
                            time.sleep(0.5)
                            continue
                else:
                    ws1 = wbxl.sheets[0]
            n_value = self.argspec.excel_value
            n_value = self.value_type(n_value)
            s_cell = self.o_ws[self.argspec.write_cell]
            s_row, s_col = s_cell.row, s_cell.column
            if self.argspec.excel_value:
                if self.argspec.range:
                    for i in range(self.min_row, self.max_row + 1):
                        for j in range(self.min_col, self.max_col + 1):
                            self.o_ws.cell(row=i+s_row-self.min_row, column=j+s_col-self.min_col).value = n_value
                else:
                    self.o_ws.cell(row=s_row, column=s_col).value = n_value
            else:
                for i in range(self.min_row, self.max_row + 1):
                    for j in range(self.min_col, self.max_col + 1):
                        if str(type(self.ws.cell(row=i, column=j))) == "<class 'openpyxl.cell.cell.Cell'>":
                            cell = self.ws.cell(row=i, column=j)
                            o_cell = self.o_ws.cell(row=i+s_row-self.min_row, column=j+s_col-self.min_col)
                            if self.argspec.data_only:
                                cl = get_column_letter(j)
                                v0 = ws1.range('%s%d' % (cl, i)).value
                                cell.value = v0
                                o_cell.value = self.value_type(cell.value)
                            else:
                                o_cell.value = self.value_type(cell.value)
        except Exception as e:
            msg = str(e)
            mcxt.logger.error(msg)
            sys.stderr.write('%s%s' % (msg, os.linesep))
        if self.argspec.data_only:
            wbxl.close()
            app.kill()
        return True

    # ==========================================================================
    def add_value(self, mcxt):
        try:
            ws1, wbxl = None, None
            if self.argspec.data_only:
                app = xlwings.App(visible=False)
                wbxl = app.books.open(self.filename)
                if self.argspec.sheet:
                    for i in range(self.argspec.sheetfind):
                        try:
                            ws1 = wbxl.sheets[self.argspec.sheet]
                            # sheet를 정상적으로 불러왔는지 확인. 간헐적으로 가져오지못함.
                            if ws1.range('A1').value:
                                break
                        except Exception as err:
                            # print(f'retry {i}')
                            mcxt.logger.error(f'get sheet retry {i}')
                            time.sleep(0.5)
                            continue
                else:
                    ws1 = wbxl.sheets[0]
            n_value = self.argspec.excel_value
            n_value = self.value_type(n_value)
            s_cell = self.o_ws[self.argspec.write_cell]
            s_row, s_col = s_cell.row, s_cell.column
            if self.argspec.excel_value:
                if self.argspec.range:
                    for i in range(self.min_row, self.max_row + 1):
                        for j in range(self.min_col, self.max_col + 1):
                            self.o_ws.cell(row=i+s_row-self.min_row, column=j+s_col-self.min_col).value += n_value
                else:
                    self.o_ws.cell(row=s_row, column=s_col).value += n_value
            else:
                for i in range(self.min_row, self.max_row + 1):
                    for j in range(self.min_col, self.max_col + 1):
                        if str(type(self.ws.cell(row=i, column=j))) == "<class 'openpyxl.cell.cell.Cell'>":
                            cell = self.ws.cell(row=i, column=j)
                            o_cell = self.o_ws.cell(row=i+s_row-self.min_row, column=j+s_col-self.min_col)
                            if self.argspec.data_only:
                                cl = get_column_letter(j)
                                v0 = ws1.range('%s%d' % (cl, i)).value
                                cell.value = v0
                                o_cell.value += self.value_type(cell.value)
                            elif self.argspec.shapes:
                                o_cell.value += self.value_type(cell.value)
                            else:
                                o_cell.value += self.value_type(cell.value)
        except Exception as e:
            msg = str(e)
            mcxt.logger.error(msg)
            sys.stderr.write('%s%s' % (msg, os.linesep))

        if self.argspec.data_only:
            wbxl.close()
            app.kill()
        return True


################################################################################
@func_log
def do_excel(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        setattr(argspec, 'formula', None)
        exl = Excel(argspec)

        if exl.extension in ('.xlsx', '.csv'):
            _ = exl.open()
            _ = exl.open_outputfile()
        elif exl.extension == '.xlsm':
            _ = exl.open(read_only=False, data_only=False, keep_vba=True)
            _ = exl.open_outputfile(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if argspec.range:
            if argspec.range.isalnum():
                exl.range_len1()
            else:
                exl.range_len3()
        exl.calc_range()
        if argspec.add:
            exl.add_value(mcxt)
        else:
            exl.overwrite_cell(mcxt)
        exl.save()
        sys.stdout.flush()
        return 0
    except OpenError as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 9
    finally:
        if exl is not None and exl.filename:
            exl.close()

        mcxt.logger.info('>>>end...')


######################
# ##########################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Simple Write',
            icon_path=get_icon_path(__file__),
            description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ##################################### for app dependent options
        # ----------------------------------------------------------------------
        mcxt.add_argument('--filename',
                          display_name='Input Excel/CSV',
                          show_default=True,
                          input_method='fileread',
                          input_group='radio=file_or_value;default',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--excel-value',
                          display_name='Input value',
                          show_default=True,
                          input_group='radio=file_or_value',
                          help='Gets the Excel value.'
                          )
        # ----------------------------------------------------------------------
        mcxt.add_argument('--sheet', '-s', nargs='?', default=None,
                          show_default=True,
                          display_name='Sheet',
                          input_group='Input',
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', '-r', default=None,
                          show_default=True,
                          display_name='Range',
                          input_group='Input',
                          help='If set read the range (eg, "A3:C9") to handle otherwise get all sheet [[A3:C9]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--write-sheet', nargs='?',
                          show_default=True,
                          display_name='Sheet',
                          input_group='Output',
                          help='Sheet name to write. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--write-cell', nargs='?', default='A1',
                          show_default=True,
                          display_name='Starting Cell',
                          const='A1',
                          input_group='Output',
                          help='Excel starting cell to write, default is [[A1]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--data-only', display_name='Data Only',
                          action='store_true',
                          help='If this flag is set get data only without formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--add', display_name='Add Value',
                          action='store_true',
                          help='Adds a value to an existing value.')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          display_name='Encoding',
                          help='Encoding for CSV file')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--sheetfind', default=5, type=int,
                          display_name='Sheet Finder',
                          help='Data Only error solution.'
                               ' Attempts to find the selected sheet the specified number of times. default is [[5]]')
        # ##################################### for app dependent parameters
        # ----------------------------------------------------------------------
        mcxt.add_argument('write', nargs='?', default=None,
                          show_default=True,
                          display_name='Output Excel/CSV',
                          input_method='filewrite',
                          help='Excel or CSV file to write, extension can be a xlsx, csv')
        argspec = mcxt.parse_args(args)
        return do_excel(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
