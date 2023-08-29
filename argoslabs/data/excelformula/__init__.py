#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excelformula`
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Formula
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2021/08/24]
#     - '--show-formula'에 type = bool -> action='store_true' 로 변경
#     - 백그라운드에서 excel이 닫히지 않는 현상 수정
#  * [2021/08/23]
#     - sheetname이 open할때 빠져있었음
#  * [2020/08/10]
#     - build a plugin
#  * [2020/08/10]
#     - starting
#

################################################################################
import os
import csv
import sys
import openpyxl
import xlwings as xw
from tempfile import gettempdir
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit

################################################################################
ENCODINGS = [
    "auto: Detecting",
    "cp932: Japanese",
    "eucjp: Japanese",
    "cp936: Chinese (simplified)",
    "cp949: Korean",
    "euckr: Korean",
    "cp950: Chinese (traditional)",
    "utf-8 UTF-8",
]


################################################################################
class Excelformula(object):

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.newcell = self.argspec.cell
        self.newfile = self.argspec.newfilename
        self.tempfile = None
        self.wb = None
        self.ws = None
        self.rr = None

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                     % os.path.basename(self.filename)[:-4])
        if self.extension == '.csv':
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
        else:
            self.wb = openpyxl.load_workbook(self.filename,
                                             read_only=read_only,
                                             data_only=False,
                                             keep_vba=keep_vba)
            if self.argspec.sheetname:
                self.ws = self.wb[self.argspec.sheetname]
            else:
                self.ws = self.wb.active
            self.ws[self.newcell] = self.argspec.formula
            self.wb.save(self.tempfile)
            self.wb.close()
        self.wb = openpyxl.load_workbook(self.tempfile,
                                         read_only=read_only,
                                         data_only=data_only,
                                         keep_vba=keep_vba)
        if self.argspec.sheetname:
            self.ws = self.wb[self.argspec.sheetname]
        else:
            self.ws = self.wb.active
        self.ws[self.newcell] = self.argspec.formula
        self.wb.save(self.tempfile)
        return self.save()

    # ==========================================================================
    def xlwingsfunc(self):
        app = xw.App(visible=False)
        try:
            if self.newfile:
                wbxl = app.books.open(self.newfile)
            else:
                wbxl = app.books.open(self.filename)
        except Exception:
            raise RuntimeError('The excel file contains the unreadable content')
        if self.argspec.sheetname:
            ws1 = wbxl.sheets[self.argspec.sheetname]
        else:
            ws1 = wbxl.sheets[0]
        v = str(ws1.range(self.newcell).value)
        try:
            v = float(v)
            v_fp = v - int(v)
            if abs(v_fp) < 0.000001:
                v = int(v)
            print(v, end='')
        except Exception:
            print(v, end='')
        wbxl.close()

    # ==========================================================================
    def A_close(self):
        if self.tempfile and os.path.exists(self.tempfile):
            os.remove(self.tempfile)
        self.wb.close()
        self.filename = None
        self.tempfile = None
        self.wb = None
        self.ws = None

    # ==========================================================================
    def xls2csv(self):
        if self.newfile:
            self.filename = self.newfile
        with open(self.filename, 'w', encoding=self.argspec.encoding) as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        return True

    # ==========================================================================
    def save(self):
        if not self.newfile:
            if self.extension == '.csv':
                self.wb.save(self.tempfile)
                self.xls2csv()
            else:
                self.wb.save(self.filename)
            t = os.path.abspath(self.filename)
            return t
        else:
            if self.extension == '.csv':
                if self.newfile.lower().endswith('csv'):
                    self.wb.save(self.tempfile)
                    self.xls2csv()
                else:
                    self.wb.save(self.newfile)
            else:
                if self.newfile.lower().endswith('csv'):
                    self.xls2csv()
                else:
                    self.wb.save(self.newfile)
            t = os.path.abspath(self.newfile)
            return t
        self.close()


################################################################################
@func_log
def do_excek2(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        exl = Excelformula(argspec)
        if exl.extension in ('.xlsx', '.csv'):
            o = exl.open()
        elif exl.extension == '.xlsm':
            o = exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if argspec.show_formula:
            exl.xlwingsfunc()
        else:
            print(o, end='')
        exl.A_close()
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        # print(False)
        return 1
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Formula',
            icon_path=get_icon_path(__file__),
            description='Read Excel Formula Value',
    ) as mcxt:
        # ######################################## for app dependent options
        mcxt.add_argument('--sheetname', display_name='Sheetname',
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='newfilename')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--show-formula', display_name='Formula Value',
                          action='store_true', help='excel formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding', default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')

        # ##################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')
        # ----------------------------------------------------------------------
        mcxt.add_argument('cell', display_name='Cell',
                          help='cell to save excel formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('formula', display_name='Formula',
                          help='excel formula')
        argspec = mcxt.parse_args(args)
        return do_excek2(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
