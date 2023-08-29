"""
====================================
 :mod:`argoslabs.data.csv2xlsx'
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for converting csv to xlsx
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2021/03/27] Kyobong An
#     - outfile path를 return 하도록 추가
#  * [2021/03/27] Jerry Chae
#     - 그룹에 "6-Files and Folders" 넣음
#  * [2021/02/23] Jerry Chae
#     - return value, Date format
# https://exceljet.net/custom-number-formats
# "$"#,##0.00_);("$"#,##0.00)
# "$"#,##0.00_);[Red]("$"#,##0.00)
# "$"#,##0.00_-
# "$"#,##0_);("$"#,##0)
# "$"#,##0_);[Red]("$"#,##0)
# # ?/?
# # ??/??
# ##0.0E+0
# #,##0
# #,##0.00
# #,##0.00_);(#,##0.00)
# #,##0.00_);[Red](#,##0.00)
# #,##0_);(#,##0)
# #,##0_);[Red](#,##0)
# $#,##0_-
# 0
# 0%
# 0.00
# 0.00%
# 0.00E+00
# @
# General
# [$EUR ]#,##0.00_-
# [h]:mm:ss
# [hh]:mm:ss
# _("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)
# _("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)
# _(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)
# _(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)
# d-m
# d-m-y
# d-mmm
# d-mmm-yy
# d/m/y
# dd/mm/yy
# h:mm
# h:mm AM/PM
# h:mm:ss
# h:mm:ss AM/PM
# h:mm:ss@
# i:s.S
# m-y
# m/d/yy h:mm
# mm-dd-yy
# mm:ss
# mmm-yy
# mmss.0
# yy-mm-dd
# yy/mm/dd@
# yyyy-mm-dd
# yyyy-mm-dd h:mm:ss
#
#  * [2020/09/23]
#     - build a plugin
#  * [2020/09/23]
#     - starting
#

################################################################################
import os
import re
import csv
import sys
import openpyxl
# import openpyxl.utils.datetime
from openpyxl.styles.numbers import BUILTIN_FORMATS, FORMAT_DATE_YYYYMMDD2, \
    FORMAT_DATE_YYMMDD, FORMAT_DATE_DDMMYY, FORMAT_DATE_DMYSLASH, FORMAT_DATE_DMYMINUS, \
    FORMAT_DATE_DMMINUS, FORMAT_DATE_MYMINUS, FORMAT_DATE_DATETIME, FORMAT_DATE_TIME7, \
    FORMAT_DATE_TIME8, FORMAT_DATE_TIMEDELTA, FORMAT_DATE_YYMMDDSLASH, \
    FORMAT_CURRENCY_USD_SIMPLE, FORMAT_CURRENCY_USD, FORMAT_CURRENCY_EUR_SIMPLE
# from openpyxl.styles import NamedStyle
# import datetime
from openpyxl.cell.cell import Cell
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit


################################################################################
def get_format_list():
    fl = [v for v in BUILTIN_FORMATS.values()]
    fl.append(FORMAT_DATE_YYYYMMDD2)
    fl.append(FORMAT_DATE_YYMMDD)
    fl.append(FORMAT_DATE_DDMMYY)
    fl.append(FORMAT_DATE_DMYSLASH)
    fl.append(FORMAT_DATE_DMYMINUS)
    fl.append(FORMAT_DATE_DMMINUS)
    fl.append(FORMAT_DATE_MYMINUS)
    fl.append(FORMAT_DATE_DATETIME)
    fl.append(FORMAT_DATE_TIME7)
    fl.append(FORMAT_DATE_TIME8)
    fl.append(FORMAT_DATE_TIMEDELTA)
    fl.append(FORMAT_DATE_YYMMDDSLASH)
    fl.append(FORMAT_CURRENCY_USD_SIMPLE)
    fl.append(FORMAT_CURRENCY_USD)
    fl.append(FORMAT_CURRENCY_EUR_SIMPLE)
    fl.sort()
    return fl


################################################################################
# noinspection PyBroadException
class csv2excelAPI(object):
    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.format = self.argspec.format
        self.opened = False
        self.wb = None
        self.ws = None
        self.outputfile = None

    # ==========================================================================
    @staticmethod
    def datelst(cell_or_tuple, fmt_):
        # DATETIME_FORMAT = {
        #     'YYYYMMDD-HHMMSS': "%Y%m%d-%H%M%S",
        #     'YYYY-MM-DD HH:MM:SS': "%Y-%m-%d %H:%M:%S",
        #     'YYYY/MM/DD HH:MM:SS': "%Y/%m/%d %H:%M:%S",
        #     'MMDDYYYY-HHMMSS': "%m%d%Y-%H%M%S",
        #     'MM-DD-YYYY HH:MM:SS': "%m-%d-%Y %H:%M:%S",
        #     'MM/DD/YYYY HH:MM:SS': "%m/%d/%Y %H:%M:%S",
        #     'M/D/YYYY HH:MM:SS': "%-m/%-d/%Y %H:%M:%S" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S",
        #     'YYYYMMDD-HHMMSS.mmm': "%Y%m%d-%H%M%S.%f",
        #     'YYYY-MM-DD HH:MM:SS.mmm': "%Y-%m-%d %H:%M:%S.%f",
        #     'YYYY/MM/DD HH:MM:SS.mmm': "%Y/%m/%d %H:%M:%S.%f",
        #     'MMDDYYYY-HHMMSS.mmm': "%m%d%Y-%H%M%S.%f",
        #     'MM-DD-YYYY HH:MM:SS.mmm': "%m-%d-%Y %H:%M:%S.%f",
        #     'MM/DD/YYYY HH:MM:SS.mmm': "%m/%d/%Y %H:%M:%S.%f",
        #     'M/D/YYYY HH:MM:SS.mmm': "%-m/%-d/%Y %H:%M:%S.%f" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S.%f",
        #     'YYYYMMDD': "%Y%m%d",
        #     'YYYY-MM-DD': "%Y-%m-%d",
        #     'YYYY/MM/DD': "%Y/%m/%d",
        #     'MMDDYYYY': "%m%d%Y",
        #     'MM-DD-YYYY': "%m-%d-%Y",
        #     'MM/DD/YYYY': "%m/%d/%Y",
        #     'M/D/YYYY': "%-m/%-d/%Y" if sys.platform != 'win32' else "%#m/%#d/%Y",
        #     'B D YYYY': "%b %-d %Y" if sys.platform != 'win32' else "%b %#d %Y",
        #     'B D, YYYY': "%b %-d, %Y" if sys.platform != 'win32' else "%b %#d, %Y",
        #     'D B YYYY': "%-d %b %Y" if sys.platform != 'win32' else "%#d %b %Y",
        # }
        # if '%' in fmt_:
        #     try:
        #         cell.value = float(cell.value.strip('%')) / 100
        #         cell.number_format = fmt_
        #     except Exception:
        #         pass
        # else:
        #     try:
        #         p = DATETIME_FORMAT[fmt_]
        #         dt = datetime.datetime.strptime(cell.value, p)
        #         cell.value = dt
        #         cell.number_format = fmt_
        #         # dst = NamedStyle()
        #     except Exception:
        #         try:
        #             cell.value = float(cell.value)
        #         except Exception:
        #             pass
        #         cell.number_format = fmt_
        if isinstance(cell_or_tuple, (tuple, list)):
            for cell in cell_or_tuple:
                if not isinstance(cell, Cell):
                    raise RuntimeError(
                        f'Invalid Cell type "{type(cell_or_tuple)}"')
                cell.number_format = fmt_
        elif isinstance(cell_or_tuple, Cell):
            cell_or_tuple.number_format = fmt_
        else:
            raise RuntimeError(f'Invalid Cell type "{type(cell_or_tuple)}"')

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        temp = os.path.splitext(self.filename.lower())[0]
        if self.argspec.newfilename:
            self.outputfile = self.argspec.newfilename
        else:
            self.outputfile = '%s.xlsx' % temp
        self.csv2xls(self.filename, self.outputfile, self.argspec.encoding)
        self.wb = openpyxl.load_workbook(self.outputfile,
                                         read_only=read_only,
                                         data_only=data_only,
                                         keep_vba=keep_vba)
        self.ws = self.wb.active
        for r in self.ws.rows:
            for i in r:
                try:
                    self.ws[i.coordinate] = float(i.value)
                except Exception:
                    pass
        self.opened = True
        return self.opened

    # ==========================================================================
    def csv2excel(self, rng, fmt):
        rn = self.ws[rng]
        if ':' in rng:
            k = rng.split(':')
            k = [i.lower() for i in k]
            if not re.match(r"\w+\d", k[0]) and k[0] == k[1]:
                for i in rn:
                    self.datelst(i, fmt)
            else:
                for i in rn:
                    for j in i:
                        self.datelst(j, fmt)
        else:
            self.datelst(rn, fmt)
        return 0

    # ==========================================================================
    def close(self):
        self.wb.save(self.outputfile)
        self.wb.close()
        return os.path.abspath(self.outputfile)


################################################################################
@func_log
def do_excelcopy(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    try:
        if argspec.list_format:
            print('\n'.join(get_format_list()), end='')
            return 0
        exl = csv2excelAPI(argspec)
        exl.open()
        if argspec.range and argspec.format:
            for i, j in zip(argspec.range, argspec.format):
                exl.csv2excel(i, j)
            print(exl.close(), end='')
        print(os.path.abspath(exl.outputfile), end='')
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    finally:
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='6',  # Files and Folders
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Csv2Xlsx',
            icon_path=get_icon_path(__file__),
            description='Convert .csv to .xlsx',
    ) as mcxt:
        # ######################################### for app dependent options
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='New filename to save ')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', display_name='Cell/Range', action='append',
                          help='Select cell or range')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--format', display_name='Cell Format',
                          choices=get_format_list(),
                          action='append', help='Cell format')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--list-format', display_name='List Format',
                          action='store_true',
                          help='If this flag is set print the list of all support cell format')

        # ###################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='CSV File',
                          input_method='fileread',
                          help='CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big'
                               ' CSV input can take time.)')
        argspec = mcxt.parse_args(args)
        return do_excelcopy(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
