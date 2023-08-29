#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excelupdate`
====================================
.. moduleauthor:: Kyobong An <akb0930@argos-labs.com>, Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel
"""
#
# Authors
# ===========
#
# * Kyobong An, Irene Cho
#
# Change Log
# --------
#  * [2021/06/14]
#   sheet와 range위치 수정
#  * [2021/05/10]
#   CSV FILE 오류수정, MergedCell 오류 수정
#  * [2021/05/03]
#   datetype 값을 변경하도록 추가, newfile 옵션추가
#  * [2021/04/30]
#   기존코드 get 삭제, repalace_value 변경 data_only사용시 원본데이터 copy해서 진행
#  * [2021/04/28]
#   셀 병합된 데이터 구분하도록 변경
#  * [2021/04/28]
#   data_only 옵션 사용시 output 변경 ex) test.xlsx -> test(1).xlsx
#  * [2021/04/26]
#   데이터 타입 비교부분 수정
#  * [2021/04/20]
#   csv파일에서 숫자인식 못하는 부분 수정
#  * [2021/04/20]
#   엑셀 오픈 부분 수정, data only시 값이 변경 안되던 부분 수정
#  * [2021/04/19]
#   코드정리, xlwings 추가
#  * [2021/03/26]
#   열엇던 파일에 바로 저장가능하도록 변경
#  * [2021/03/19]
#   단일행 오류 수정
#  * [2021/03/18]
#   행과열 전체 읽기 추가
#  * [2021/03/17]
#   코드정리
################################################################################
import os
import sys
import csv
# import time
import datetime
import shutil
import openpyxl
import xlwings
from tempfile import gettempdir
from openpyxl.utils import get_column_letter
from io import StringIO
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import ModuleContext, func_log, \
    ArgsError, ArgsExit, get_icon_path
import warnings

warnings.filterwarnings("ignore")

DATETIME_FORMAT = {
    'YYYYMMDD-HHMMSS': "%Y%m%d-%H%M%S",
    'YYYY-MM-DD HH:MM:SS': "%Y-%m-%d %H:%M:%S",
    'YYYY/MM/DD HH:MM:SS': "%Y/%m/%d %H:%M:%S",
    'MMDDYYYY-HHMMSS': "%m%d%Y-%H%M%S",
    'MM-DD-YYYY HH:MM:SS': "%m-%d-%Y %H:%M:%S",
    'MM/DD/YYYY HH:MM:SS': "%m/%d/%Y %H:%M:%S",
    'M/D/YYYY HH:MM:SS': "%-m/%-d/%Y %H:%M:%S" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S",
    'YYYYMMDD-HHMMSS.mmm': "%Y%m%d-%H%M%S.%f",
    'YYYY-MM-DD HH:MM:SS.mmm': "%Y-%m-%d %H:%M:%S.%f",
    'YYYY/MM/DD HH:MM:SS.mmm': "%Y/%m/%d %H:%M:%S.%f",
    'MMDDYYYY-HHMMSS.mmm': "%m%d%Y-%H%M%S.%f",
    'MM-DD-YYYY HH:MM:SS.mmm': "%m-%d-%Y %H:%M:%S.%f",
    'MM/DD/YYYY HH:MM:SS.mmm': "%m/%d/%Y %H:%M:%S.%f",
    'M/D/YYYY HH:MM:SS.mmm': "%-m/%-d/%Y %H:%M:%S.%f" if sys.platform != 'win32' else "%#m/%#d/%Y %H:%M:%S.%f",
    'YYYYMMDD': "%Y%m%d",
    'YYYY-MM-DD': "%Y-%m-%d",
    'YYYY/MM/DD': "%Y/%m/%d",
    'MMDDYYYY': "%m%d%Y",
    'MM-DD-YYYY': "%m-%d-%Y",
    'MM/DD/YYYY': "%m/%d/%Y",
    'M/D/YYYY': "%-m/%-d/%Y" if sys.platform != 'win32' else "%#m/%#d/%Y",
    'B D YYYY': "%b %-d %Y" if sys.platform != 'win32' else "%b %#d %Y",
    'B D, YYYY': "%b %-d, %Y" if sys.platform != 'win32' else "%b %#d, %Y",
    'D B YYYY': "%-d %b %Y" if sys.platform != 'win32' else "%#d %b %Y",
}


################################################################################
class OpenError(Exception):
    pass


################################################################################
class ExtensionError(Exception):
    pass


################################################################################
class Excel(object):
    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel file "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.wb = None
        self.ws = None
        self.min_row = self.max_row = -1
        self.min_col = self.max_col = -1
        self.rr = None
        self.rrt = None
        # for csv reading
        self.tempfile = None
        self.temp_cell = ''

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def __repr__(self):
        sio = StringIO()
        wr = csv.writer(sio, lineterminator='\n')
        for row in self.rr:
            wr.writerow(row)
        return sio.getvalue()

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
            f.close()
            wb.close()
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def xls2csv(self):
        with open(self.filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        self.extension = self._get_extension(self.filename)
        if self.argspec.data_only and not self.argspec.newfile:
            t = self._get_safe_next_filename(self.filename)
            shutil.copy(self.filename, t)
            self.filename = t
        elif self.argspec.newfile:
            try:
                shutil.copy(self.filename, self.argspec.newfile)
            except Exception:
                raise ExtensionError(
                    f'The extension "{self.argspec.newfile}" does not match the original extension.')

            self.filename = self.argspec.newfile
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.load_workbook(self.filename,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if not self.argspec.sheet:
                self.ws = self.wb.active
            elif self.argspec.sheet in self.wb.sheetnames:
                self.ws = self.wb[self.argspec.sheet]
            else:
                raise OpenError(
                    f'The sheet name "{self.argspec.sheetname}" does not exist')
        return True

    # ==========================================================================
    def close(self):
        if self.open:
            self.wb.close()
            if self.tempfile and os.path.exists(self.tempfile):
                os.remove(self.tempfile)
            self.ws = None
            self.wb = None
            return True
        return False

    # ==========================================================================
    def range_len1(self):       # 단일 행이나 열을 체크 ex) 'A'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha():
            self.rr[0] = self.rr[0] + '1'
        elif self.rr[0].isdigit():
            self.rr[0] = 'A' + self.rr[0]
        else:
            self.rr = self.rr * 2

    def range_len3(self):      # 두줄이상의 행이나 열체크  ex) 'A:B','2:16'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha() and self.rr[1].isalpha():
            self.rr[0] = self.rr[0] + '1'
            self.rr[1] = self.rr[1] + '1'
        elif self.rr[0].isdigit() and self.rr[1].isdigit():
            self.rr[0] = 'A' + self.rr[0]
            self.rr[1] = 'A' + self.rr[1]
        else:
            pass

    # ==========================================================================
    def calc_range(self):
        if not self.open:
            return False
        self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
        self.min_col, self.max_col = self.ws.min_column, self.ws.max_column
        # 범위가 None 값일때 전체범위
        if not self.argspec.range:
            self.temp_cell = 'A1'
            return False

        self.rrt = self.argspec.range.strip().split(':')

        self.temp_cell = self.rr[0]
        if self.argspec.range.isalnum():    # 단일행이나 열의 볌위 'A'
            if self.argspec.range.isalpha():
                c = self.ws[self.rr[0]]
                self.min_col, self.max_col = c.column, c.column
                return True
            elif self.argspec.range.isdigit():
                c = self.ws[self.rr[0]]
                self.min_row, self.max_row = c.row, c.row
                return True

        if self.rrt[0].isalpha() and self.rrt[1].isalpha():    # 범위  'A:B'  처럼 열범위
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_col, self.max_col = min(c1.column, c2.column), max(c1.column, c2.column)

        elif self.rrt[0].isdigit() and self.rrt[1].isdigit():   # 범위 '1:16' 처럼 행범위체크
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row, c2.row)

        else:                  # 특정범위를 지정  'A4:E4'
            if self.argspec.range.isalnum():
                c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            else:
                r1, r2 = self.argspec.range.strip().split(':')
                c1, c2 = self.ws[r1.strip()], self.ws[r2.strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row, c2.row)
            self.min_col, self.max_col = min(c1.column, c2.column), max(c1.column, c2.column)
        return True

    # ==========================================================================
    def save(self):
        argspec_write = self.filename
        if self.extension == '.csv':    # 만약 결과를 csv에 쓴다면
            self.wb.save(self.tempfile)
            self.xls2csv()
            print(os.path.abspath(argspec_write), end='')
            self.wb.close()
        else:   # 나머지는 모두 excel 파일이라 생각
            self.wb.save(argspec_write)
            print(os.path.abspath(argspec_write), end='')
            self.wb.close()
        return 0

    # ==========================================================================
    @staticmethod
    def value_type(value, fmt_=None):
        if not value:
            value = None
            return value
        elif isinstance(value, datetime.datetime):
            try:
                p = DATETIME_FORMAT[fmt_]
                value = value.strftime(p)
            except Exception:
                pass
        else:
            try:
                value = float(value)
            except Exception as e:
                t = str(e).split(':')[1].strip()
                if "ufeff3" in t:
                    raise OpenError(f'Cannot read the value {t}. Check the encoding of the file.')
                pass
        return value

    # ==========================================================================
    def replace_cell(self):
        ws1, wbxl = None, None
        if self.argspec.data_only:
            app = xlwings.App(visible=False)
            wbxl = app.books.open(self.filename)
            if self.argspec.sheet:
                ws1 = wbxl.sheets[self.argspec.sheet]
            else:
                ws1 = wbxl.sheets[0]
        c_value, n_value = self.argspec.excel_value, self.argspec.replace_value
        c_value, n_value = self.value_type(c_value), self.value_type(n_value)
        for i in range(self.min_row, self.max_row + 1):
            for j in range(self.min_col, self.max_col + 1):
                if str(type(self.ws.cell(row=i, column=j))) == "<class 'openpyxl.cell.cell.Cell'>":
                    cell = self.ws.cell(row=i, column=j)
                    if self.argspec.data_only:
                        cl = get_column_letter(j)
                        v0 = ws1.range('%s%d' % (cl, i)).value
                        cell.value = v0
                    v = self.value_type(cell.value, self.argspec.format)
                    if c_value == v:
                        cell.value = n_value
        if self.argspec.data_only:
            wbxl.close()
        return True


##############################################################################
@func_log
def do_excel(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        setattr(argspec, 'formula', None)
        exl = Excel(argspec)
        if exl.extension in ('.xlsx', '.csv'):
            _ = exl.open()
        elif exl.extension == '.xlsm':
            _ = exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        if argspec.range:
            if argspec.range.isalnum():
                exl.range_len1()
            else:
                exl.range_len3()
        exl.calc_range()
        exl.replace_cell()
        exl.save()
        sys.stdout.flush()
        return 0
    except OpenError as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    except ExtensionError as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 2
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 9
    finally:
        if exl:
            exl.close()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    """
    Build user argument and options and call plugin job function
    :param args: user arguments
    :return: return value from plugin job function
    """
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Update',
            icon_path=get_icon_path(__file__),
            description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ##################################### for app dependent options
        mcxt.add_argument('--sheet', '-s', nargs='?',
                          display_name='Read-fr sheet', show_default=True,
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--range', '-r', show_default=True,
                          display_name='Read range',
                          help='If set read the range (eg, "A3:C9") to handle otherwise get all sheet [[A3:C9]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--format', display_name='Format',
                          default='YYYY-MM-DD',
                          choices=list(DATETIME_FORMAT.keys()),
                          help='Specify cell value format e.g. YYYY-MM-DD')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--data-only', display_name='Data Only',
                          action='store_true',
                          help='If this flag is set get data only without formula')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          display_name='Encoding',
                          help='Encoding for CSV file')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfile',
                          display_name='New file',
                          input_method='filewrite',
                          help="It updates the new file without modifying the original file."
                          "Don't list the extensions.")
        # ##################################### for app dependent parameters
        # ----------------------------------------------------------------------
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')
        # ----------------------------------------------------------------------
        mcxt.add_argument('excel_value',
                          display_name='Current Value',
                          help='The value you want to change, default = None')
        # ----------------------------------------------------------------------
        mcxt.add_argument('replace_value',
                          display_name='New Value',
                          help='Value to be changed')
        argspec = mcxt.parse_args(args)
        return do_excel(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
