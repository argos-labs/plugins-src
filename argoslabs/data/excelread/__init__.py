#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.excelread`
====================================
.. moduleauthor:: Kyobong An <akb0930@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel
"""
#
# Authors
# ===========
#
# * Kyobong An
#
# Change Log
# --------
#  * [2022/07/12]
#   - 엑셀 읽는 방식 변경 기존에는 행과열로 이중 포문 사용 하여 제어. ->  iter_rows를 사용해 최대최소 행과열을 변수로하는 함수 사용.
#   - 읽는 속도 개선. 하지만 위와같은 방식은 openpyxl에서만 적용 dataonly를 사용하는 xlwings에서는 사용불가.
#  * [2022/02/03]
#   - .xlsx과 .csv 형식의 파일만이 아닌 다른 파일들도 열수 있도록 변경.
#   - file open 할때, read_only=True로 변경함. 일부 파일에 접근할 수 없음. ex)Slicer List
#  * [2021/06/14]
#   숫자에서 뒤에 .0 붙는부분 수정, range와 sheet 위치 수정
#  * [2021/04/26]
#  데이터 타입 비교부분 수정
#  * [2021/04/16]
#   xlwings 추가, open error추가
################################################################################
import os
import sys
import csv
import time
import shutil
# noinspection PyPackageRequirements
import openpyxl
import xlwings
from tempfile import gettempdir
# noinspection PyPackageRequirements
from openpyxl.utils import get_column_letter
from io import StringIO
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import ModuleContext, func_log, \
    ArgsError, ArgsExit, get_icon_path
import warnings
warnings.filterwarnings("ignore")


################################################################################
class OpenError(Exception):
    pass


################################################################################
class Excel(object):
    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel file "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.opened = False
        self.wb = None
        self.ws = None
        self.min_row = self.max_row = -1
        self.min_col = self.max_col = -1
        self.rr = None
        self.rrt = None
        # for csv reading
        self.tempfile = None
        self.temp_cell = ''

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    def __repr__(self):
        sio = StringIO()
        wr = csv.writer(sio, lineterminator='\n')
        for row in self.rr:
            wr.writerow(row)
        return sio.getvalue()

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
            f.close()
            wb.close()
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    def xls2csv(self):
        with open(self.filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        return True

    # ==========================================================================
    def open(self, read_only=True, data_only=False, keep_vba=False):
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.load_workbook(self.filename,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if self.argspec.sheet is None:
                self.ws = self.wb.active
            elif self.argspec.sheet in self.wb.sheetnames:
                self.ws = self.wb[self.argspec.sheet]
            else:
                msg = 'Sheet name does not exist'
                raise OpenError(msg)
                # self.ws = self.wb.create_sheet(self.argspec.sheet)
        self.opened = True
        return self.opened

    # ==========================================================================
    def save_formula(self):
        if not (self.opened and self.argspec.formula):
            return False
        for fp in self.argspec.formula:
            fr, fm = fp.split('=', maxsplit=1)
            self.ws[fr].value = '=%s' % fm
        # 윈도우에서는 열었던 것을 저장할 수 없다함. 다른 파일로 임시로 저장하고
        # renmae 시켜봄
        tmp_fn = '%s-tmp.xlsx' % self.filename
        self.wb.save(tmp_fn)
        self.close()
        shutil.move(tmp_fn, self.filename)
        self.open(read_only=True)
        return True

    # ==========================================================================
    def close(self):
        if self.opened:
            self.wb.close()
            if self.tempfile and os.path.exists(self.tempfile):
                # wb를 close 후에 삭제하여야 함
                for _ in range(3):
                    # noinspection PyBroadException
                    try:
                        os.remove(self.tempfile)
                        break
                    except Exception:
                        time.sleep(0.3)
            self.ws = None
            self.wb = None
            self.opened = False
            return True
        return False

    # ==========================================================================
    def range_len1(self):       # 단일 행이나 열을 체크 ex) 'A'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha():
            self.rr[0] = self.rr[0] + '1'
        elif self.rr[0].isdigit():
            self.rr[0] = 'A' + self.rr[0]
        else:
            self.rr = self.rr * 2

    def range_len3(self):      # 두줄이상의 행이나 열체크  ex) 'A:B','2:16'
        self.rr = self.argspec.range.strip().split(':')
        if self.rr[0].isalpha() and self.rr[1].isalpha():
            self.rr[0] = self.rr[0] + '1'
            self.rr[1] = self.rr[1] + '1'
        elif self.rr[0].isdigit() and self.rr[1].isdigit():
            self.rr[0] = 'A' + self.rr[0]
            self.rr[1] = 'A' + self.rr[1]
        else:
            pass

    # ==========================================================================
    def calc_range(self):
        if not self.opened:
            return False
        if not self.argspec.range:          # 범위가 None 값일때 전체범위
            self.temp_cell = 'A1'
            self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
            self.min_col, self.max_col = self.ws.min_column, self.ws.max_column
            return False

        self.rrt = self.argspec.range.strip().split(':')

        self.temp_cell = self.rr[0]
        if self.argspec.range.isalnum():    # 단일행이나 열의 볌위 'A'
            if self.argspec.range.isalpha():
                c = self.ws[self.rr[0]]
                self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
                self.min_col, self.max_col = c.column, c.column
                return True
            elif self.argspec.range.isdigit():
                c = self.ws[self.rr[0]]
                self.min_row, self.max_row = c.row, c.row
                self.min_col, self.max_col = self.ws.min_column, self.ws.max_column
                return True

        if self.rrt[0].isalpha() and self.rrt[1].isalpha():    # 범위  'A:B'  처럼 열범위
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_row, self.max_row = self.ws.min_row, self.ws.max_row
            self.min_col, self.max_col = min(c1.column, c2.column), max(c1.column, c2.column)

        elif self.rrt[0].isdigit() and self.rrt[1].isdigit():   # 범위 '1:16' 처럼 행범위체크
            c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row, c2.row)
            self.min_col, self.max_col = self.ws.min_column, self.ws.max_column

        else:                  # 특정범위를 지정  'A4:E4'
            if self.argspec.range.isalnum():
                c1, c2 = self.ws[self.rr[0].strip()], self.ws[self.rr[1].strip()]
            else:
                r1, r2 = self.argspec.range.strip().split(':')
                c1, c2 = self.ws[r1.strip()], self.ws[r2.strip()]
            self.min_row, self.max_row = min(c1.row, c2.row), max(c1.row, c2.row)
            self.min_col, self.max_col = min(c1.column, c2.column), max(c1.column, c2.column)
        return True

    # ==========================================================================
    def _get(self):
        ws1, wbxl = None, None
        if self.argspec.data_only:
            app = xlwings.App(visible=False)
            wbxl = app.books.open(self.filename)
            if self.argspec.sheet:
                ws1 = wbxl.sheets[self.argspec.sheet]
            else:
                ws1 = wbxl.sheets[0]
        if self.argspec.data_only:
            for r in range(self.min_row, self.max_row + 1):
                row = []
                for c in range(self.min_col, self.max_col + 1):
                    cl = get_column_letter(c)
                    if str(type(ws1['%s%d' % (cl, r)].value)) == "<class 'float'>":
                        v = ws1['%s%d' % (cl, r)].value
                        if v - int(v) == 0:
                            v = int(v)
                    else:
                        v = ws1['%s%d' % (cl, r)].value

                    if v is None:
                        v = ''
                    row.append(v)
                yield row
        else:
            for r in self.ws.iter_rows(self.min_row, self.max_row, self.min_col, self.max_col):
                row = []
                for cell in r:
                    v = cell.value
                    if v is None:
                        v = ''
                    row.append(v)
                yield row
        if self.argspec.data_only:
            wbxl.close()

    # ==========================================================================
    def get(self):
        self.rr = []

        # not big option
        for row in self._get():
            self.rr.append(row)

        # s = str(self)
        s = str(self).strip()  # 셀 하나인 경우에도 개행문자가 붙는 것 제거

        sys.stdout.write(s)
        return bool(self.rr)

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn, overwrite=False):
        if overwrite:
            return fn
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    def Classification(self):
        for i in range(self.min_row, self.max_row + 1):
            for j in range(self.min_col, self.max_col + 1):
                if not self.ws.cell(row=i, column=j).value:
                    pass
                elif str(type(self.ws.cell(row=i, column=j).value)) == "<class 'int'>":
                    self.ws.cell(row=i, column=j).value = int(self.ws.cell(row=i, column=j).value)
                elif str(type(self.ws.cell(row=i, column=j).value)) == "<class 'float'>":
                    self.ws.cell(row=i, column=j).value = float(self.ws.cell(row=i, column=j).value)
                else:
                    pass


################################################################################
@func_log
def do_excel(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        setattr(argspec, 'formula', None)
        exl = Excel(argspec)
        data_only = argspec.data_only
        # 기존에 .xlsx 과 .csv만 열도록 했는데 다른 파일형식도 열수 있도록 변경.
        # if exl.extension in ('.xlsx', '.csv', '.xlsm'):
        exl.open(data_only=data_only)

        if argspec.range is not None:
            if '' == argspec.range:
                argspec.range = None

        if argspec.range is not None:
            if argspec.range.isalnum():
                exl.range_len1()
            else:
                exl.range_len3()

        exl.calc_range()
        # exl.Classification()
        exl.get()
        # exl.save()
        sys.stdout.flush()
        return 0
    except OpenError as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 9
    finally:
        if exl is not None:
            exl.close()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    """
    Build user argument and options and call plugin job function
    :param args: user arguments
    :return: return value from plugin job function
    """
    with ModuleContext(
        owner='ARGOS-LABS',
        group='2',
        version='1.0',
        platform=['windows', 'darwin', 'linux'],
        output_type='csv',
        display_name='Excel Simple Read',
        icon_path=get_icon_path(__file__),
        description='Excel reading, func call and optional writing',
    ) as mcxt:
        # ##################################### for app dependent options
        mcxt.add_argument('--sheet', '-s', nargs='?', default=None,
                          display_name='Read sheet', show_default=True,
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        mcxt.add_argument('--range', default=None,
                          display_name='Read range', show_default=True,
                          help='If set read the range (eg, "A3:C9") to handle otherwise get all sheet [[A3:C9]]')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--data-only', display_name='Data Only',
                          action='store_true',
                          help='If this flag is set get data only without formula')
        # # -----------------------------------------------------------------+-----
        mcxt.add_argument('--encoding',
                          display_name='Encoding',
                          help='Encoding for CSV file')

        # ##################################### for app dependent parameters
        # ----------------------------------------------------------------------
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big CSV input can take time.)')

        argspec = mcxt.parse_args(args)
        return do_excel(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
