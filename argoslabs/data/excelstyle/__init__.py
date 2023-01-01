"""
====================================
 :mod:`argoslabs.data.excelstyle`
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Style
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2021/11/18] TAIKI.S
#     - "Use functions" 기능 추가. "Compare Values"에 함수 추가 할 수 있도록 해주는 기능.
#  * [2021/07/13]
#     - build a plugin
#  * [2021/07/13]
#     - starting
#

################################################################################
import os
import re
import csv
import sys
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import FormulaRule, CellIsRule, Rule
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit


################################################################################
class Excelformula(object):

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.newfile = self.argspec.newfilename
        self.tempfile = None
        self.wb = None
        self.ws = None

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        self.wb = openpyxl.load_workbook(self.filename,
                                         read_only=read_only,
                                         data_only=data_only,
                                         keep_vba=keep_vba)
        if self.argspec.sheetname is None:
            self.ws = self.wb.active
        elif self.argspec.sheetname:
            self.ws = self.wb[self.argspec.sheetname]

    # ==========================================================================
    def calc_range(self, rng):
        if ':' in rng:
            k = rng.split(':')
            k = [i.lower() for i in k]
            if not re.match("\w+\d", k[0]) and k[0] == k[1]:
                return 'same_col'  # A:A
            else:
                return 'diff_col'
        else:
            if self.ws[rng]:
                return 'cell'
            else:
                raise RuntimeError('Cannot read the range %s' % rng)

    # ==========================================================================
    def stylefunc(self):
        for ent in self.argspec.range:
            if self.argspec.operator:
                color = PatternFill(bgColor=self.argspec.fillcolor)
                if self.argspec.functions:
                    self.ws.conditional_formatting.add(
                        ent, Rule(type="expression", formula=[self.argspec.value],
                                  dxf=DifferentialStyle(fill=color))
                    )
                else:
                    self.ws.conditional_formatting.add(
                        ent, CellIsRule(operator=self.argspec.operator, formula=[self.argspec.value],
                                         fill=color))
            # print(ent)
            rng_type = self.calc_range(ent)
            t = self.ws[ent]
            if rng_type == 'same_col':
                for i in t:
                    self.ws[i.coordinate].font = Font(bold=self.argspec.bold,
                                                      italic=self.argspec.italic,
                                                      underline=self.argspec.underline)
            elif rng_type == 'diff_col':
                for i in t:
                    for x in i:
                        self.ws[x.coordinate].font = Font(
                            bold=self.argspec.bold, italic=self.argspec.italic,
                            underline=self.argspec.underline)
            else:
                t.font = Font(bold=self.argspec.bold,
                              italic=self.argspec.italic,
                              underline=self.argspec.underline)

        print(self.save(), end='')

    # ==========================================================================
    def save(self):
        if not self.newfile:
            self.wb.save(self.filename)
            t = os.path.abspath(self.filename)
            return t
        else:
            self.wb.save(self.newfile)
        t = os.path.abspath(self.newfile)
        self.wb.close()
        return t


################################################################################
@func_log
def do_excek2(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    try:
        exl = Excelformula(argspec)
        if exl.extension in '.xlsx':
            exl.open()
        else:
            raise RuntimeError(
                'Only support file extension of "*.xlsx"')
        exl.stylefunc()
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Style',
            icon_path=get_icon_path(__file__),
            description='Change excel font style',
    ) as mcxt:
        # ######################################## for app dependent options
        mcxt.add_argument('--sheetname', display_name='Sheetname',
                          help='Sheet name to handle. If not specified last activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--bold', display_name='Bold', action='store_true',
                          default=False, help='Bold font')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--italic', display_name='Italic',
                          action='store_true',
                          default=False, help='Italic fond')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--underline', display_name='Underline', default=None,
                          choices=['double', 'single'], help='Underline')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--fillcolor', display_name='Fill Color',
                          help='Select color to fill the cell')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--operator', display_name='Operators',
                          choices=['>','>=','<','<=','=','!='],
                          input_group='Fill Color Conditions',
                          help='Conditions to fill color')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--value', display_name='Compare Values',
                          input_group='Fill Color Conditions',
                          help='Compare values')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--functions', display_name='Use functions',
                          input_group='Fill Color Conditions',
                          action='store_true',
                          default=False,
                          help='Using function in conditional formatting rule')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='New filename', input_method='filewrite')
        # ##################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel File',
                          input_method='fileread',
                          help='Excel filename to handle for reading')
        # ----------------------------------------------------------------------
        mcxt.add_argument('range', display_name='Cell/Range', nargs='+',
                          help='Cell or range in excel file')
        argspec = mcxt.parse_args(args)
        return do_excek2(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
