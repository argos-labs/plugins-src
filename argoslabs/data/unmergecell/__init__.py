#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`argoslabs.data.json.unmergecell'
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
ARGOS LABS plugin module for Excel Unmerge Cell
"""
#
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#
#  * [2020/10/12]
#     - build a plugin
#  * [2020/10/12]
#     - starting
#

################################################################################
import os
import csv
import sys
import shutil
import openpyxl
from tempfile import gettempdir
from openpyxl.utils import range_boundaries
from alabs.common.util.vvencoding import get_file_encoding
from alabs.common.util.vvargs import func_log, get_icon_path, ModuleContext, \
    ArgsError, ArgsExit


################################################################################
# noinspection PyBroadException
class ExcelunmergeAPI(object):
    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.filename = argspec.filename
        if not os.path.exists(self.filename):
            raise IOError('Cannot read excel filename "%s"' % self.filename)
        self.extension = self._get_extension(self.filename)
        self.opened = False
        self.wb = None
        self.ws = None
        self.tempfile = None

    # ==========================================================================
    @staticmethod
    def _get_extension(fn):
        sps = os.path.splitext(fn.lower())
        return sps[-1]

    # ==========================================================================
    @staticmethod
    def csv2xls(s, t, encoding=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        if not encoding:
            encoding = get_file_encoding(s)
        try:
            with open(s, encoding=encoding) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(t)
        except Exception as e:
            raise RuntimeError('Cannot read CSV file. Please check encoding: %s'
                               % str(e))
        return True

    # ==========================================================================
    @staticmethod
    def _get_safe_next_filename(fn):
        fn, ext = os.path.splitext(fn)
        for n in range(1, 1000000):
            nfn = f'{fn} ({n})' + ext
            if not os.path.exists(nfn):
                return nfn

    # ==========================================================================
    def xls2csv(self):
        if self.argspec.newfilename:
            self.filename = self.argspec.newfilename
        with open(self.filename, 'w', encoding='utf-8') as f:
            c = csv.writer(f, lineterminator='\n')
            for r in self.ws.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        return True

    # ==========================================================================
    def open(self, read_only=False, data_only=False, keep_vba=False):
        if self.argspec.dataonly:
            data_only = True
            if not self.argspec.newfilename:
                t = self._get_safe_next_filename(self.filename)
                shutil.copy(self.filename, t)
                self.filename = t
        if self.extension == '.csv':
            self.tempfile = os.path.join(gettempdir(), '%s.xlsx'
                                         % os.path.basename(self.filename)[:-4])
            self.csv2xls(self.filename, self.tempfile, self.argspec.encoding)
            self.wb = openpyxl.load_workbook(self.tempfile,
                                             read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            self.ws = self.wb.active
            for r in self.ws.rows:
                for i in r:
                    try:
                        self.ws[i.coordinate] = float(i.value)
                    except Exception:
                        pass
        else:
            self.wb = openpyxl.load_workbook(self.filename, read_only=read_only,
                                             data_only=data_only,
                                             keep_vba=keep_vba)
            if self.argspec.sheetname is None:
                self.ws = self.wb.active
            elif self.argspec.sheetname:
                self.ws = self.wb[self.argspec.sheetname]
        self.opened = True
        return self.opened

    # ==========================================================================
    def save(self):
        if not self.argspec.newfilename:
            if self.extension == '.csv':
                self.wb.save(self.tempfile)
                self.xls2csv()
                self.close()
            else:
                self.wb.save(self.filename)
            return os.path.abspath(self.filename)
        else:
            if self.extension == '.csv':
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.wb.save(self.tempfile)
                    self.xls2csv()
                    self.close()
                else:
                    self.wb.save(self.argspec.newfilename)
            else:
                if self.argspec.newfilename.lower().endswith('csv'):
                    self.xls2csv()
                else:
                    self.wb.save(self.argspec.newfilename)
                self.wb.close()
            return os.path.abspath(self.argspec.newfilename)

    # ==========================================================================
    def close(self):
        if self.tempfile and os.path.exists(self.tempfile):
            os.remove(self.tempfile)

    # ==========================================================================
    def unmerge(self):
        t = self.ws.merged_cells.ranges
        lst = []
        for i in t:
            lst.append(i)
        for i in lst:
            print(t)
            min_col, min_row, max_col, max_row = range_boundaries(str(i))
            top_left_cell_value = self.ws.cell(row=min_row,
                                               column=min_col).value
            self.ws.unmerge_cells(str(i))
            for row in self.ws.iter_rows(min_col=min_col, min_row=min_row,
                                         max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
        return 0


################################################################################
@func_log
def do_excelunmerge(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    exl = None
    try:
        exl = ExcelunmergeAPI(argspec)
        if exl.extension in ('.xlsx', '.csv', '.xls'):
            exl.open()
        elif exl.extension == '.xlsm':
            exl.open(read_only=False, data_only=False, keep_vba=True)
        else:
            raise RuntimeError(
                'Only support file extension of ("*.csv", "*.xlsx", "*.xlsm")')
        exl.unmerge()
        print(exl.save(), end='')
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        # print(False)
        return 1
    finally:
        if exl is not None:
            exl.close()
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='csv',
            display_name='Excel Unmerge Cells',
            icon_path=get_icon_path(__file__),
            description='Excel unmerge cells',
    ) as mcxt:
        # ######################################### for app dependent options
        mcxt.add_argument('--sheetname', display_name='Copy from Sheet',
                          help='Sheet name to handle. If not specified last'
                               ' activated sheet is chosen')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--newfilename', display_name='Save As',
                          help='newfilename')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--dataonly', display_name='Data Only',
                          action='store_true',
                          help='print only data.json not formulas')
        # ----------------------------------------------------------------------
        mcxt.add_argument('--encoding',
                          default='utf-8',
                          display_name='Encoding',
                          help='Encoding for CSV file, default is [[utf-8]]')

        # ###################################### for app dependent parameters
        mcxt.add_argument('filename',
                          display_name='Excel/CSV File',
                          input_method='fileread',
                          help='Excel or CSV filename to handle for reading. '
                               '(Note. CSV is converted to excel first. Too big'
                               ' CSV input can take time.)')
        argspec = mcxt.parse_args(args)
        return do_excelunmerge(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
