"""
====================================
 :mod:`argoslabs.google.googlesheet`
====================================
.. moduleauthor:: Irene Cho <irene@argos-labs.com>
.. note:: ARGOS-LABS License

Description
===========
Manage Google Sheets
"""
# Authors
# ===========
#
# * Irene Cho
#
# Change Log
# --------
#  * [2022/03/07] Kyobong An
#     - 뒤쪽 열에 값이 없을 경우 none으로 인식이 안됨 첫번째 행의 길이를 아래 행에도 첫번째행의 길이만큼 ''를 삽입하도록함.
#  * [2020/06/19]
#     - Chnage output format
#
#  * [2020/06/15]
#     - starting

################################################################################
from __future__ import print_function
import os
import csv
import sys
import pickle
import os.path
import openpyxl
from io import StringIO
from googleapiclient.discovery import build
from alabs.common.util.vvargs import ModuleContext, func_log, \
    ArgsError, ArgsExit, get_icon_path
import warnings


################################################################################
class Googlesheet(object):
    # ==========================================================================
    OP_TYPE = ['Read a Spreadsheet', 'Create a Spreadsheet',
               'Write a Spreadsheet', 'Rename a Spreadsheet',
               'Add a Sheet', 'Delete a Sheet', 'Rename a Sheet',
               'Duplicate a Sheet', 'Find and Replace']

    # ==========================================================================
    def __init__(self, argspec):
        self.argspec = argspec
        self.service = None
        self.spreadsheet_id = self.argspec.spreadsheet_id
        self.range = self.argspec.range
        self.csvfile = self.argspec.csvfile
        self.requests = None
        self.result = None

    # ==========================================================================
    def read_file(self):
        if not self.argspec.token:
            raise IOError('Cannot read token file "%s"' % self.argspec.token)
        k = self.argspec.token
        with open(k, 'rb') as token:
            creds = pickle.load(token)
        self.service = build('sheets', 'v4', credentials=creds)
        return self.service

    # ==========================================================================
    def callsheet(self):
        sheet = self.service.spreadsheets()
        result = sheet.values().get(spreadsheetId=self.spreadsheet_id,
                                    range=self.range).execute()
        values = result.get('values', [])
        s = csv.writer(sys.stdout, lineterminator='\n')
        for i in values:
            if len(values[0]) != len(i):
                for _ in range(len(values[0]) - len(i)):
                    i.append('')
            s.writerow(i)

    # ==========================================================================
    def create_sheet(self):
        spreadsheet = {
            'properties': {
                'title': self.argspec.title

            }
        }
        spreadsheet = self.service.spreadsheets().create(body=spreadsheet,
                                                         fields='spreadsheetId').execute()
        self.result = {}
        print(spreadsheet.get('spreadsheetId'), end="")

    # ==========================================================================
    def opencsv(self):
        if not os.path.exists(self.csvfile):
            raise IOError('Cannot read csv file "%s"' % self.csvfile)
        else:
            k = os.path.splitext(self.csvfile.lower())[-1]
            if k == '.csv':
                with open(self.csvfile, encoding='utf-8') as f:
                    reader = csv.reader(f)
                    data = list(reader)
            elif k == '.xlsx':
                wb = openpyxl.load_workbook(self.csvfile)
                if self.argspec.xlsxsheetname:
                    ws = wb[self.argspec.xlsxsheetname]
                else:
                    ws = wb.active
                data = []
                for r in ws.rows:
                    data.append([cell.value for cell in r])
                wb.close()
            return data

    # ==========================================================================
    def write_sheet(self):
        values = self.opencsv()
        body = {'values': values}
        self.result = self.service.spreadsheets().values().update(
            spreadsheetId=self.spreadsheet_id, range=self.range,
            valueInputOption='USER_ENTERED', body=body).execute()
        print(self.spreadsheet_id, end='')

    # ==========================================================================
    def add_sheet(self):
        self.requests = [
            {'addSheet': {'properties': {"sheetId": self.argspec.sheetid,
                                         "title": self.argspec.sheetitle}}}]

    # ==========================================================================
    def delete_sheet(self):
        self.requests = [{'deleteSheet': {"sheetId": self.argspec.sheetid}}]

    # ==========================================================================
    def rename_sheet(self):
        self.requests = [{'updateSheetProperties': {'properties': {
            "sheetId": self.argspec.sheetid, "title": self.argspec.sheetitle},
            'fields': 'title'}}]

    # ==========================================================================
    def change_title(self):
        self.requests = [{'updateSpreadsheetProperties': {
            'properties': {'title': self.argspec.title}, 'fields': 'title'}}]
        self.result = self.service.spreadsheets().batchUpdate(
            spreadsheetId=self.spreadsheet_id,
            body={'requests': self.requests}).execute()
        print(self.spreadsheet_id, end='')

    # ==========================================================================
    def duplicate_sheet(self):
        self.requests = [{'duplicateSheet': {
            'sourceSheetId': self.argspec.sheetid,  # default is 0
            "insertSheetIndex": self.argspec.index,
            "newSheetName": self.argspec.sheetitle}}]

    # ==========================================================================
    def findreplace(self):
        self.requests = [{
            'findReplace': {'find': self.argspec.find,
                            'replacement': self.argspec.replace}}]
        if self.argspec.allSheets:
            self.requests[0]['findReplace']['allSheets'] = True
            self.argspec.sheetid = -1
        elif self.argspec.sheetid:
            self.requests[0]['findReplace']['sheetId'] = self.argspec.sheetid
        else:
            raise IOError('Cannot find Sheet Id')

    # ==========================================================================
    def print(self):
        dict0 = {}
        if self.argspec.op == self.OP_TYPE[4]:
            dict0 = self.result['replies'][0]['addSheet']['properties']
            dict0 = {k: v for k, v in dict0.items() if k.startswith('sheetId')}
        elif self.argspec.op == self.OP_TYPE[7]:
            temp = self.result['replies'][0]['duplicateSheet']['properties']
            if temp['sheetId']:
                dict0['new sheetId'] = temp['sheetId']
            dict0['old sheetId'] = self.argspec.sheetid
            self.argspec.sheetid = None
        elif self.argspec.op == self.OP_TYPE[8]:
            dict0 = self.result['replies'][0]['findReplace']
            dict0 = {k: v for k, v in dict0.items() if k.startswith('sheetId')}
        if self.argspec.sheetid:
            dict0['sheetId'] = self.argspec.sheetid
        dict0['spreadsheetId'] = self.result['spreadsheetId']
        with StringIO() as outst:
            dict_writer = csv.DictWriter(outst, dict0.keys())
            dict_writer.writeheader()
            dict_writer.writerows([dict0])
            print(outst.getvalue(), end='')

    # ==========================================================================
    def do(self):
        if self.argspec.op == self.OP_TYPE[0]:
            self.callsheet()
        elif self.argspec.op == self.OP_TYPE[1]:
            self.create_sheet()
        elif self.argspec.op == self.OP_TYPE[2]:
            self.write_sheet()
        elif self.argspec.op == self.OP_TYPE[3]:
            self.change_title()
        else:
            if self.argspec.op == self.OP_TYPE[4]:
                self.add_sheet()
            elif self.argspec.op == self.OP_TYPE[5]:
                self.delete_sheet()
            elif self.argspec.op == self.OP_TYPE[6]:
                self.rename_sheet()
            elif self.argspec.op == self.OP_TYPE[7]:
                self.duplicate_sheet()
            elif self.argspec.op == self.OP_TYPE[8]:
                self.findreplace()
            self.result = self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body={'requests': self.requests}).execute()
            self.print()


################################################################################
@func_log
def reg_op(mcxt, argspec):
    mcxt.logger.info('>>>starting...')
    try:
        warnings.simplefilter("ignore", ResourceWarning)
        res = Googlesheet(argspec)
        res.read_file()
        res.do()
        return 0
    except Exception as err:
        msg = str(err)
        mcxt.logger.error(msg)
        sys.stderr.write('%s%s' % (msg, os.linesep))
        return 1
    finally:
        sys.stdout.flush()
        mcxt.logger.info('>>>end...')


################################################################################
def _main(*args):
    with ModuleContext(
            owner='ARGOS-LABS-DEMO',
            group='2',
            version='1.0',
            platform=['windows', 'darwin', 'linux'],
            output_type='text',
            display_name='Google Sheets',
            icon_path=get_icon_path(__file__),
            description='Managing Google Sheets',
    ) as mcxt:
        # ##################################### for app dependent parameters
        mcxt.add_argument('op', display_name='Function type',
                          choices=Googlesheet.OP_TYPE,
                          help='Type of operation')
        mcxt.add_argument('token', display_name='Token.Pickle',
                          help='Google Sheet API Token',
                          input_method='fileread')
        # ##################################### for app dependent parameters
        mcxt.add_argument('--spreadsheet_id', display_name='Spreadsheet Id',
                          help='Spreadsheet Id')
        mcxt.add_argument('--sheetid', display_name='Sheet Id',
                          help='Sheet Id')
        mcxt.add_argument('--range', display_name='Sheet Range',
                          help='Sheet Range')
        mcxt.add_argument('--title', display_name='Spreadsheet Name',
                          help='New Spreadsheet Name')
        mcxt.add_argument('--sheetitle', display_name='Sheet Name',
                          help='New Sheet Name')
        mcxt.add_argument('--csvfile', display_name='Excel/CSVfile',
                          help='csvfile for writing values',
                          input_method='fileread')
        mcxt.add_argument('--xlsxsheetname', display_name='Inputfile Sheet',
                          help='Excel Sheet Name')
        mcxt.add_argument('--index', display_name='Specify Index', default=0,
                          type=int,
                          help='specify the index for the duplicated sheet')
        mcxt.add_argument('--find', display_name='Find',
                          help='find values in Google Sheet')
        mcxt.add_argument('--replace', display_name='Replace',
                          help='replace values in Google Sheet')
        mcxt.add_argument('--allSheets', display_name='AllSheets',
                          help='replace all values in the spreadsheet',
                          action='store_true', default=False)
        argspec = mcxt.parse_args(args)
        return reg_op(mcxt, argspec)


################################################################################
def main(*args):
    try:
        return _main(*args)
    except ArgsError as err:
        sys.stderr.write('Error: %s\nPlease -h to print help\n' % str(err))
    except ArgsExit as _:
        pass
